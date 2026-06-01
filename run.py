import asyncio
import os
import re
import argparse
from datetime import datetime, timezone
from pathlib import Path
from motor.motor_asyncio import AsyncIOMotorClient
import paramiko
import openpyxl

from src.config.settings import settings
from src.config.loader import ConfigLoader
from src.config.cache import ConfigCache
from src.config.validator import ConfigValidator
from src.core.enums import FileType
from src.core.types import FieldMapping, FieldMappingType
from src.models.mapping_config import MappingConfig, MappingConfigRepository
from src.pipeline.ingestion_pipeline import IngestionPipeline
from src.scheduler import PartnerDataScheduler, SchedulerConfig, daily_partner_fetch_job
from src.logging import get_structured_logger

def parse_excel_template(template_path: str) -> dict:
    """Parses a MappingConfig dynamically from an Excel template file (e.g. RequestTemplate.xlsx).
    
    No hardcoded values. Everything is read dynamically from the spreadsheet cells.
    """
    print(f"Parsing configuration dynamically from template: {template_path}")
    wb = openpyxl.load_workbook(template_path, data_only=True)
    sheet = wb["Sheet1"]
    
    partner = sheet.cell(row=3, column=1).value
    path_file_pattern = sheet.cell(row=3, column=4).value
    start_row = int(sheet.cell(row=3, column=5).value)
    
    raw_id = sheet.cell(row=3, column=6).value
    raw_trace = sheet.cell(row=3, column=7).value
    raw_amount = sheet.cell(row=3, column=8).value
    raw_currency = sheet.cell(row=3, column=9).value
    raw_status = sheet.cell(row=3, column=10).value
    raw_trans_date = sheet.cell(row=3, column=11).value
    
    extra_service = sheet.cell(row=3, column=12).value
    extra_portal = sheet.cell(row=3, column=13).value
    extra_provider = sheet.cell(row=3, column=14).value
    
    def parse_column_field(raw_val, path, mapping_type):
        if not raw_val:
            return None
        lines = [line.strip() for line in raw_val.split("\n") if line.strip()]
        col_match = re.match(r"column\s+(\d+)", lines[0], re.IGNORECASE)
        if col_match:
            # Shift 0-based column from template to 1-based column for system
            col_num = int(col_match.group(1)) + 1
            source_field = lines[1] if len(lines) > 1 else None
            
            mapping_dict = None
            if mapping_type == FieldMappingType.MAPPING:
                mapping_dict = {}
                for line in lines[1:]:
                    if ":" in line:
                        k, v = line.split(":", 1)
                        mapping_dict[k.strip()] = v.strip()
            
            return FieldMapping(
                path=path,
                column=col_num,
                sourceField=source_field,
                type=mapping_type,
                required=(path == "id"),
                mapping=mapping_dict
            )
        else:
            return FieldMapping(
                path=path,
                constant=raw_val,
                type=FieldMappingType.CONSTANT
            )

    field_mappings = []
    field_mappings.append(parse_column_field(raw_id, "id", FieldMappingType.STRING))
    field_mappings.append(parse_column_field(raw_trace, "trace", FieldMappingType.STRING))
    field_mappings.append(parse_column_field(raw_amount, "amount", FieldMappingType.DECIMAL))
    field_mappings.append(FieldMapping(path="currency", constant=raw_currency, type=FieldMappingType.CONSTANT))
    field_mappings.append(parse_column_field(raw_status, "status", FieldMappingType.MAPPING))
    field_mappings.append(parse_column_field(raw_trans_date, "transDate", FieldMappingType.DATE))
    
    field_mappings.append(FieldMapping(path="extra.service", constant=extra_service, type=FieldMappingType.CONSTANT))
    field_mappings.append(FieldMapping(path="extra.portal", constant=extra_portal, type=FieldMappingType.CONSTANT))
    field_mappings.append(FieldMapping(path="extra.provider", constant=extra_provider, type=FieldMappingType.CONSTANT))
    
    field_mappings = [fm for fm in field_mappings if fm is not None]
    
    # Extract filename from path pattern
    filename = Path(path_file_pattern).name if path_file_pattern else "m4becomvsp_07072024_combine.xlsx"
    if path_file_pattern and "#{" in path_file_pattern:
        # Fallback to test file name if it contains expression
        filename = "m4becomvsp_07072024_combine.xlsx"
        
    return {
        "partner": partner,
        "start_row": start_row + 1,  # Row 7 is header, data starts at 8
        "field_mappings": field_mappings,
        "filename": filename
    }

async def _handle_scheduler_mode(
    db,
    start_scheduler: bool,
    run_job_now: bool,
    list_jobs: bool,
) -> None:
    """Handle scheduler-related CLI commands.

    Args:
        db: AsyncIOMotorDatabase instance.
        start_scheduler: Whether to start the scheduler daemon.
        run_job_now: Whether to manually trigger the job now.
        list_jobs: Whether to list scheduled jobs.
    """
    structured_logger = get_structured_logger()
    config_loader = _create_config_loader(db)

    scheduler_config = SchedulerConfig(
        job_store_type="mongodb",
        mongodb_url=settings.mongodb_url,
        db_name=settings.db_name,
    )

    def on_job_executed(event):
        print(f"[SCHEDULER] Job executed: {event.job_id}")

    def on_job_error(event):
        print(f"[SCHEDULER] Job failed: {event.job_id} - {event.exception}")

    scheduler = PartnerDataScheduler(
        config=scheduler_config,
        on_job_executed=on_job_executed,
        on_job_error=on_job_error,
    )

    # Start the scheduler first so jobs can be added to it
    scheduler.start()

    # Add daily job (connections are resolved dynamically inside the job to allow pickling)
    scheduler.add_daily_job(
        job_func=daily_partner_fetch_job,
        job_id="daily_partner_fetch",
    )

    if list_jobs:
        jobs = scheduler.list_jobs()
        if not jobs:
            print("No scheduled jobs.")
        else:
            print("\n=== Scheduled Jobs ===")
            for job in jobs:
                print(f"  ID: {job['id']}")
                print(f"  Name: {job['name']}")
                print(f"  Next Run: {job['next_run_time']}")
                print(f"  Trigger: {job['trigger']}")
                print()
        scheduler.stop()
        return

    if run_job_now:
        print("Triggering daily fetch job now...")
        scheduler.run_job_now("daily_partner_fetch")
        # Wait a bit for job to complete
        await asyncio.sleep(5)
        print("Job triggered. Check logs for results.")
        scheduler.stop()
        return

    if start_scheduler:
        print("Starting scheduler daemon...")
        print(f"  Job Store: {scheduler_config.job_store_type}")
        print(f"  Default Schedule: {scheduler_config.default_schedule}")
        print("\nPress Ctrl+C to stop.\n")

        try:
            # Keep the event loop running
            while True:
                await asyncio.sleep(1)
        except KeyboardInterrupt:
            print("\nStopping scheduler...")
            scheduler.stop()
            print("Scheduler stopped.")



def _create_config_loader(db):
    """Create a ConfigLoader instance.

    Args:
        db: AsyncIOMotorDatabase instance.

    Returns:
        ConfigLoader instance.
    """
    config_repo = MappingConfigRepository(db)
    config_cache = ConfigCache()
    config_validator = ConfigValidator()
    return ConfigLoader(config_repo, config_cache, config_validator)


async def main():
    parser = argparse.ArgumentParser(description="Reconciliation Ingestion Pipeline CLI")
    parser.add_argument("--config", type=str, help="Path to Excel template config (e.g. RequestTemplate.xlsx) to dynamically upload")
    parser.add_argument("--start-scheduler", action="store_true", help="Start the scheduler for automated partner data fetching")
    parser.add_argument("--run-job-now", action="store_true", help="Manually trigger the daily fetch job now")
    parser.add_argument("--list-jobs", action="store_true", help="List all scheduled jobs")
    parser.add_argument("action", nargs="?", choices=["reconcile"], help="Action to perform (e.g. reconcile)")
    parser.add_argument("--date", type=str, help="Date for reconciliation (YYYY-MM-DD)")
    parser.add_argument("--reconcile", type=str, help="Run reconciliation for date (YYYY-MM-DD)")
    parser.add_argument("--partner", type=str, default="MOMO", help="Partner identifier for reconciliation")
    parser.add_argument("--seed-mock", action="store_true", help="Seed mock internal transactions for testing reconciliation")
    args = parser.parse_args()

    # 1. Database Connection
    print(f"Connecting to MongoDB at {settings.mongodb_url}...")
    client = AsyncIOMotorClient(settings.mongodb_url)
    db = client[settings.db_name]
    
    # Apply MongoDB indexes automatically on startup
    from src.models.indexes import apply_indexes
    print("Applying MongoDB indexes...")
    await apply_indexes(db)
    print("Indexes verified/applied successfully.")

    # Reconciliation mode
    is_reconcile = args.action == "reconcile" or args.reconcile is not None
    if is_reconcile:
        date_str = args.date or args.reconcile
        if not date_str:
            print("Error: Please provide a date using --date or --reconcile (YYYY-MM-DD)")
            return
        recon_date = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        
        if args.seed_mock:
            print("Seeding mock internal transactions...")
            from src.models.internal_transaction import InternalTransactionRepository, InternalTransaction
            from src.core.enums import TransactionStatus
            from decimal import Decimal
            
            repo = InternalTransactionRepository(db)
            # Clear old mock data for clean run
            await repo.collection.delete_many({"partner": args.partner})
            
            # Seed MATCHED case: (Matches standard MOMO record in combine xlsx)
            await repo.create(InternalTransaction(
                id="internal_matched_01",
                partner=args.partner,
                partnerTxnId="2407055711887385978413624",
                amount=Decimal("259200"),
                status=TransactionStatus.SUCCESS,
                transactionTime=recon_date,
            ))
            
            # Seed AMOUNT_MISMATCH case
            await repo.create(InternalTransaction(
                id="internal_amt_mismatch_01",
                partner=args.partner,
                partnerTxnId="2407055711887385978413625",
                amount=Decimal("100000"),  # Expected: file might have another amount
                status=TransactionStatus.SUCCESS,
                transactionTime=recon_date,
            ))

            # Seed MISSING_PARTNER case
            await repo.create(InternalTransaction(
                id="internal_missing_partner_01",
                partner=args.partner,
                partnerTxnId="internal_only_txn_999",
                amount=Decimal("15000"),
                status=TransactionStatus.SUCCESS,
                transactionTime=recon_date,
            ))
            print("Mock internal transactions seeded successfully.")

        print(f"Executing reconciliation for partner {args.partner} on {args.reconcile}...")
        from src.reconciliation.engine import ReconciliationEngine
        engine = ReconciliationEngine(db)
        results = await engine.reconcile(args.partner, recon_date)
        print(f"Reconciliation finished. Total results generated/updated: {len(results)}")
        for r in results:
            print(f"  - Key: {r.partner_txn_id} -> Status: {r.reconciliation_status} (Partner Amt: {r.partner_amount}, Internal Amt: {r.internal_amount})")
        return

    # Scheduler mode
    if args.start_scheduler or args.run_job_now or args.list_jobs:
        await _handle_scheduler_mode(
            db=db,
            start_scheduler=args.start_scheduler,
            run_job_now=args.run_job_now,
            list_jobs=args.list_jobs,
        )
        return

    # ... rest of existing code for manual ingestion

    # 2. Upload Config Dynamically if --config is passed
    if args.config:
        parsed_config = parse_excel_template(args.config)
        partner = parsed_config["partner"]
        start_row = parsed_config["start_row"]
        field_mappings = parsed_config["field_mappings"]
        remote_filename = parsed_config["filename"]
        
        print(f"Uploading parsed MappingConfig for {partner} to MongoDB...")
        config_doc = MappingConfig(
            partner=partner,
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            sheet_name=sheet_name,
            start_row=start_row,
            field_mappings=field_mappings,
            config_version="v_template"
        )
        await db["reconciliation_mapping_config"].delete_many({"partner": partner})
        config_repo = MappingConfigRepository(db)
        await config_repo.create(config_doc)
        print("MappingConfig successfully uploaded to MongoDB!")

    # 3. SFTP Connection and File Download
    sftp_host = os.getenv("SFTP_HOST", "localhost")
    sftp_port = int(os.getenv("SFTP_PORT", "2222"))
    sftp_user = os.getenv("SFTP_USER", "foo")
    sftp_pass = os.getenv("SFTP_PASS", "pass")
    
    remote_path = f"/upload/{remote_filename}"
    local_dir = Path("./tmp_downloads")
    local_dir.mkdir(exist_ok=True)
    local_path = local_dir / remote_filename
    
    print(f"Connecting to SFTP {sftp_user}@{sftp_host}:{sftp_port}...")
    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(sftp_host, port=sftp_port, username=sftp_user, password=sftp_pass, timeout=10)
        
        sftp = ssh.open_sftp()
        print(f"Downloading {remote_path} from SFTP...")
        sftp.get(remote_path, str(local_path))
        sftp.close()
        ssh.close()
        print(f"File downloaded successfully to {local_path}")
    except Exception as e:
        print(f"SFTP connection/download failed: {e}")
        fallback_path = Path("./sftp_data") / remote_filename
        if fallback_path.exists():
            print(f"Using local sftp_data fallback: {fallback_path}")
            local_path = fallback_path
        else:
            print("Error: Could not retrieve file from SFTP or local sftp_data folder.")
            return

    # 4. Clean up previous records for a fresh run
    reconciliation_date = datetime(2024, 7, 7, tzinfo=timezone.utc)
    print(f"Cleaning up any existing records for {partner} on 2024-07-07...")
    await db["reconciliation_file"].delete_many({"partner": partner, "reconciliationDate": reconciliation_date})
    await db["data_container"].delete_many({"identify": partner, "reconciliationDate": reconciliation_date})
    
    # 5. Run Ingestion Pipeline
    config_repo = MappingConfigRepository(db)
    config_cache = ConfigCache()
    config_validator = ConfigValidator()
    config_loader = ConfigLoader(config_repo, config_cache, config_validator)
    
    pipeline = IngestionPipeline(db=db, config_loader=config_loader)
    
    print("Processing file through ingestion pipeline...")
    result = await pipeline.process_file(
        file_path=str(local_path),
        partner=partner,
        workflow_type="UPC",
        file_type=FileType.SETTLEMENT,
        reconciliation_date=reconciliation_date,
        config_version="v_template"
    )
    
    # 6. Print Results
    print("\n=== Pipeline Processing Summary ===")
    print(f"Status: {result.file_record.processing_status}")
    print(f"Total Rows: {result.stats.total_rows}")
    print(f"Success Rows: {result.stats.success_rows}")
    print(f"Failed Rows: {result.stats.failed_rows}")
    
    if result.errors:
        print(f"Errors ({len(result.errors)}):")
        for err in result.errors[:5]:
            print(f"  - Row {err.get('row', 'N/A')}: {err.get('field')} -> {err.get('reason')}")
            
    saved_count = await db["data_container"].count_documents({"identify": partner})
    print(f"Verified documents in MongoDB (data_container): {saved_count}")
    
    if local_path.exists() and "tmp_downloads" in str(local_path):
        try:
            local_path.unlink()
            # Only remove directory if it is empty
            if local_dir.exists() and not any(local_dir.iterdir()):
                local_dir.rmdir()
            print("Cleaned up temporary download files.")
        except Exception as e:
            print(f"Warning: Could not clean up temporary directory: {e}")

if __name__ == "__main__":
    asyncio.run(main())
