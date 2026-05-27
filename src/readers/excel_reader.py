"""Excel streaming reader using openpyxl in read-only mode."""

from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path
from typing import Self

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.models.mapping_config import MappingConfig

VALID_EXTENSIONS = {".xlsx", ".xlsm"}


class ExcelStreamReader:
    """Memory-efficient Excel file reader using openpyxl read-only mode.

    Opens Excel files in read-only mode for constant memory usage regardless
    of file size. Supports context manager protocol for automatic resource cleanup.
    """

    # Default patterns for summary/footer row detection (case-insensitive)
    DEFAULT_SKIP_PATTERNS: list[str] = [
        "total",
        "grand total",
        "summary",
        "footer",
        "合计",
        "总计",
        "小计",
    ]

    def __init__(
        self,
        file_path: str | Path,
        *,
        sheet_name: str | None = None,
        sheet_index: int | None = None,
        start_row: int = 1,
        skip_empty_rows: bool = True,
        skip_patterns: list[str] | None = None,
    ) -> None:
        """Initialize the reader with a file path.

        Args:
            file_path: Path to the Excel file (.xlsx or .xlsm).
            sheet_name: Select sheet by name (mutually exclusive with sheet_index).
            sheet_index: Select sheet by 0-based index (mutually exclusive with sheet_name).
            start_row: 1-based row number to start reading from (default 1 = first row).
            skip_empty_rows: Skip rows where all cells are None or empty string (default True).
            skip_patterns: List of string patterns; skip rows where ANY cell contains
                a pattern match (case-insensitive). Defaults to common summary/footer keywords.

        Raises:
            FileNotFoundError: If the file does not exist.
            ValueError: If the file extension is not .xlsx or .xlsm.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        if path.suffix.lower() not in VALID_EXTENSIONS:
            raise ValueError(
                f"Unsupported file extension '{path.suffix}'. "
                f"Must be one of: {', '.join(sorted(VALID_EXTENSIONS))}"
            )
        self._file_path: Path = path
        self._workbook: Workbook | None = None
        self._worksheet: Worksheet | None = None
        self._sheet_name: str | None = sheet_name
        self._sheet_index: int | None = sheet_index
        self._start_row: int = start_row
        self._skip_empty_rows: bool = skip_empty_rows
        self._skip_patterns: list[str] | None = (
            skip_patterns if skip_patterns is not None
            else self.DEFAULT_SKIP_PATTERNS
        )

    @classmethod
    def from_mapping_config(
        cls, file_path: str | Path, config: MappingConfig
    ) -> ExcelStreamReader:
        """Create an ExcelStreamReader from a MappingConfig object.

        Uses config.sheet_name and config.start_row to configure the reader.
        Empty row skipping is enabled by default.

        Args:
            file_path: Path to the Excel file (.xlsx or .xlsm).
            config: MappingConfig containing sheet_name and start_row.

        Returns:
            Configured ExcelStreamReader instance.
        """
        return cls(
            file_path=file_path,
            sheet_name=config.sheet_name,
            start_row=config.start_row,
            skip_empty_rows=True,
        )

    def __enter__(self) -> Self:
        """Load the workbook in read-only mode and select the worksheet.

        Returns:
            self for use within the context manager.
        """
        self._workbook = load_workbook(
            filename=str(self._file_path), read_only=True
        )
        self._select_sheet()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> bool:
        """Close the workbook, even if an exception occurred.

        Returns:
            False — does not suppress exceptions.
        """
        if self._workbook is not None:
            self._workbook.close()
            self._workbook = None
        return False

    def _require_workbook(self) -> None:
        """Raise RuntimeError if workbook is not loaded.

        Raises:
            RuntimeError: If called outside a context manager block.
        """
        if self._workbook is None:
            raise RuntimeError("Reader must be used as context manager")

    def get_sheet_names(self) -> list[str]:
        """Return the list of sheet names in the workbook.

        Returns:
            List of sheet name strings.

        Raises:
            RuntimeError: If called outside a context manager block.
        """
        self._require_workbook()
        return self._workbook.sheetnames  # type: ignore[union-attr]

    def _select_sheet(self) -> None:
        """Select the worksheet based on sheet_name, sheet_index, or default.

        Raises:
            KeyError: If sheet_name is provided but not found.
            IndexError: If sheet_index is provided but out of range.
        """
        self._require_workbook()
        if self._sheet_name is not None:
            if self._sheet_name not in self._workbook.sheetnames:  # type: ignore[union-attr]
                raise KeyError(f"Sheet '{self._sheet_name}' not found")
            self._worksheet = self._workbook[self._sheet_name]  # type: ignore[index]
        elif self._sheet_index is not None:
            try:
                self._worksheet = self._workbook.worksheets[self._sheet_index]  # type: ignore[union-attr]
            except IndexError:
                raise IndexError(
                    f"Sheet index {self._sheet_index} out of range "
                    f"(workbook has {len(self._workbook.worksheets)} sheets)"  # type: ignore[arg-type]
                )
        else:
            self._worksheet = self._workbook.active  # type: ignore[assignment]

    @staticmethod
    def _is_empty_row(row: tuple) -> bool:
        """Check if a row is empty (all cells are None or empty string).

        Args:
            row: Tuple of cell values from openpyxl.

        Returns:
            True if ALL cells are None or empty string, False otherwise.
        """
        return all(cell is None or cell == "" for cell in row)

    def _should_skip_row(self, row: tuple) -> bool:
        """Determine whether a row should be skipped.

        Checks empty rows (if skip_empty_rows is True) and pattern matches
        (if skip_patterns is provided).

        Args:
            row: Tuple of cell values from openpyxl.

        Returns:
            True if the row should be skipped, False otherwise.
        """
        if self._skip_empty_rows and self._is_empty_row(row):
            return True

        if self._skip_patterns:
            for cell in row:
                if cell is None:
                    continue
                cell_str = str(cell).lower()
                for pattern in self._skip_patterns:
                    if pattern.lower() in cell_str:
                        return True

        return False

    def iter_rows(self) -> Iterator[tuple]:
        """Iterate over rows in the selected worksheet starting from start_row.

        Rows are filtered: empty rows (all None/empty) are skipped when
        skip_empty_rows is True (default). Rows matching skip_patterns
        (case-insensitive substring match) are also skipped.

        Yields:
            Tuples of cell values (str, int, float, datetime, None, etc.).
            Empty cells yield None.

        Raises:
            RuntimeError: If called outside a context manager block.
        """
        self._require_workbook()
        for row in self._worksheet.iter_rows(  # type: ignore[union-attr]
            min_row=self._start_row, values_only=True
        ):
            if self._should_skip_row(row):
                continue
            yield row
