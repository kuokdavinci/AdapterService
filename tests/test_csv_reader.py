"""Test suite for CSVStreamReader and dynamic reader selection."""

import tempfile
from pathlib import Path
from typing import Generator

import pytest
from openpyxl import Workbook

from src.core.enums import FileType
from src.models.mapping_config import MappingConfig
from src.readers import create_reader
from src.readers.csv_reader import CSVStreamReader
from src.readers.excel_reader import ExcelStreamReader


@pytest.fixture
def simple_csv_file() -> Generator[Path, None, None]:
    """Create a simple CSV file with headers and data."""
    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
        path = Path(f.name)
    path.write_text("Header1,Header2\nA,B\nC,D\nE,F\n", encoding="utf-8")
    yield path
    path.unlink(missing_ok=True)


@pytest.fixture
def empty_rows_csv_file() -> Generator[Path, None, None]:
    """Create a CSV file with empty rows interspersed."""
    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
        path = Path(f.name)
    path.write_text("Header1,Header2\nA,B\n\nC,D\n,\nE,F\n", encoding="utf-8")
    yield path
    path.unlink(missing_ok=True)


@pytest.fixture
def summary_rows_csv_file() -> Generator[Path, None, None]:
    """Create a CSV file with summary/footer rows."""
    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
        path = Path(f.name)
    path.write_text(
        "Description,Amount\nA,100\nTotal,100\nGrand Total,100\nNormal,200\n",
        encoding="utf-8",
    )
    yield path
    path.unlink(missing_ok=True)


def _temp_path(suffix: str) -> Path:
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as f:
        return Path(f.name)


def _mapping_config(start_row: int = 1) -> MappingConfig:
    return MappingConfig(
        partner="test",
        workflowType="recon",
        fileType=FileType.SETTLEMENT,
        sheetName="IgnoredForCsv",
        startRow=start_row,
        fieldMappings=[],
    )


class TestCSVStreamReaderInit:
    """Tests for CSVStreamReader initialization and validation."""

    def test_file_not_found_raises(self) -> None:
        missing = _temp_path(".csv")
        missing.unlink()
        with pytest.raises(FileNotFoundError, match="File not found"):
            CSVStreamReader(missing)

    def test_non_csv_extension_raises(self) -> None:
        txt_file = _temp_path(".txt")
        txt_file.write_text("a,b\n1,2", encoding="utf-8")
        try:
            with pytest.raises(ValueError, match="Unsupported file extension"):
                CSVStreamReader(txt_file)
        finally:
            txt_file.unlink(missing_ok=True)

    def test_path_object_accepted(self, simple_csv_file: Path) -> None:
        reader = CSVStreamReader(simple_csv_file)
        assert reader._file_path == simple_csv_file

    def test_string_path_accepted(self, simple_csv_file: Path) -> None:
        reader = CSVStreamReader(str(simple_csv_file))
        assert reader._file_path == simple_csv_file


class TestCSVRowIteration:
    """Tests for CSV row iteration behavior."""

    def test_basic_iteration(self, simple_csv_file: Path) -> None:
        with CSVStreamReader(
            simple_csv_file,
            skip_empty_rows=False,
            skip_patterns=[],
        ) as reader:
            rows = list(reader.iter_rows())

        assert len(rows) == 4
        assert rows[0] == ("Header1", "Header2")
        assert rows[1] == ("A", "B")

    def test_start_row_offset(self, simple_csv_file: Path) -> None:
        with CSVStreamReader(
            simple_csv_file,
            start_row=2,
            skip_empty_rows=False,
            skip_patterns=[],
        ) as reader:
            rows = list(reader.iter_rows())

        assert rows == [("A", "B"), ("C", "D"), ("E", "F")]

    def test_iter_rows_outside_context_raises(self, simple_csv_file: Path) -> None:
        reader = CSVStreamReader(simple_csv_file)
        with pytest.raises(RuntimeError, match="context manager"):
            list(reader.iter_rows())

    def test_empty_rows_skipped_by_default(self, empty_rows_csv_file: Path) -> None:
        with CSVStreamReader(empty_rows_csv_file) as reader:
            rows = list(reader.iter_rows())

        assert len(rows) == 4
        assert ("", "") not in rows
        assert () not in rows

    def test_empty_rows_kept_when_disabled(self, empty_rows_csv_file: Path) -> None:
        with CSVStreamReader(
            empty_rows_csv_file,
            skip_empty_rows=False,
            skip_patterns=[],
        ) as reader:
            rows = list(reader.iter_rows())

        assert len(rows) == 6
        assert () in rows
        assert ("", "") in rows

    def test_skip_patterns_matches_case_insensitive(
        self, summary_rows_csv_file: Path
    ) -> None:
        with CSVStreamReader(
            summary_rows_csv_file,
            skip_empty_rows=False,
            skip_patterns=["total"],
        ) as reader:
            rows = list(reader.iter_rows())

        row_values = [r[0] for r in rows]
        assert "Total" not in row_values
        assert "Grand Total" not in row_values
        assert "Normal" in row_values

    def test_custom_delimiter(self) -> None:
        path = _temp_path(".csv")
        path.write_text("id;amount\nTXN001;100000\n", encoding="utf-8")

        try:
            with CSVStreamReader(
                path,
                delimiter=";",
                skip_empty_rows=False,
                skip_patterns=[],
            ) as reader:
                rows = list(reader.iter_rows())
        finally:
            path.unlink(missing_ok=True)

        assert rows == [("id", "amount"), ("TXN001", "100000")]

    def test_utf8_encoding(self) -> None:
        path = _temp_path(".csv")
        path.write_text("status\nThành công\n", encoding="utf-8")

        try:
            with CSVStreamReader(
                path,
                encoding="utf-8",
                skip_empty_rows=False,
                skip_patterns=[],
            ) as reader:
                rows = list(reader.iter_rows())
        finally:
            path.unlink(missing_ok=True)

        assert rows == [("status",), ("Thành công",)]


class TestCSVMappingConfigIntegration:
    """Tests for from_mapping_config factory method."""

    def test_from_mapping_config_uses_start_row(self, simple_csv_file: Path) -> None:
        reader = CSVStreamReader.from_mapping_config(
            simple_csv_file,
            _mapping_config(start_row=2),
        )

        assert reader._start_row == 2

    def test_from_mapping_config_enables_empty_row_skip(
        self, simple_csv_file: Path
    ) -> None:
        reader = CSVStreamReader.from_mapping_config(
            simple_csv_file,
            _mapping_config(),
        )

        assert reader._skip_empty_rows is True

    def test_from_mapping_config_returns_correct_type(
        self, simple_csv_file: Path
    ) -> None:
        reader = CSVStreamReader.from_mapping_config(
            simple_csv_file,
            _mapping_config(),
        )

        assert isinstance(reader, CSVStreamReader)


class TestCSVContextManager:
    """Tests for context manager behavior and cleanup."""

    def test_context_manager_opens_file(self, simple_csv_file: Path) -> None:
        reader = CSVStreamReader(simple_csv_file)
        assert reader._file is None
        with reader:
            assert reader._file is not None

    def test_context_manager_closes_file(self, simple_csv_file: Path) -> None:
        with CSVStreamReader(simple_csv_file) as reader:
            assert reader._file is not None
        assert reader._file is None

    def test_context_manager_closes_on_exception(self, simple_csv_file: Path) -> None:
        reader = CSVStreamReader(simple_csv_file)
        with pytest.raises(ValueError):
            with reader:
                raise ValueError("test error")
        assert reader._file is None


class TestCreateReader:
    """Tests for dynamic reader selection."""

    def test_csv_extension_returns_csv_reader(self, simple_csv_file: Path) -> None:
        reader = create_reader(simple_csv_file, _mapping_config())
        assert isinstance(reader, CSVStreamReader)

    def test_xlsx_extension_returns_excel_reader(self) -> None:
        path = _temp_path(".xlsx")
        wb = Workbook()
        wb.active.append(["A"])
        wb.save(str(path))
        wb.close()

        try:
            reader = create_reader(path, _mapping_config())
        finally:
            path.unlink(missing_ok=True)
        assert isinstance(reader, ExcelStreamReader)

    def test_xlsm_extension_returns_excel_reader(self) -> None:
        path = _temp_path(".xlsm")
        wb = Workbook()
        wb.active.append(["A"])
        wb.save(str(path))
        wb.close()

        try:
            reader = create_reader(path, _mapping_config())
        finally:
            path.unlink(missing_ok=True)
        assert isinstance(reader, ExcelStreamReader)

    def test_unsupported_extension_raises(self) -> None:
        path = _temp_path(".txt")
        path.write_text("a,b\n", encoding="utf-8")

        try:
            with pytest.raises(ValueError, match="Unsupported file extension"):
                create_reader(path, _mapping_config())
        finally:
            path.unlink(missing_ok=True)
