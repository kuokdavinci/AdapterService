"""Integration tests for IngestionPipeline — full pipeline scenarios.

Tests cover:
1. Happy path — 10 valid rows, all persisted, stats correct
2. Mixed valid/invalid — 7 valid, 3 invalid rows, stats correct, errors collected
3. All invalid rows — 0 success, all failed, status COMPLETED
4. Empty file (header only) — total=0, status COMPLETED
5. Duplicate file hash — early return with error, no processing
6. Batch insertion — 250 rows, verify insert_many called 3 times (100 + 100 + 50)
7. Exception during processing — status set to FAILED, partial stats returned
8. DataContainer records have correct source_file_id linking to ReconciliationFile
9. Partner data correctly nested in DataContainer (amount, currency, status, trace)
10. ProcessingStats returned matches ReconciliationFile stats

All tests use mocked repositories — no MongoDB connection required.
"""

import tempfile
import uuid
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from src.core.enums import FileType, ProcessingStatus, TransactionStatus
from src.core.types import (
    FieldMapping,
    FieldMappingType,
    ProcessingStats,
)
from src.models.data_container import DataContainer, DataContainerRepository, PartnerData
from src.models.mapping_config import MappingConfig
from src.models.reconciliation_file import (
    ReconciliationFile,
    ReconciliationFileRepository,
)
from src.pipeline import IngestionPipeline, IngestionResult


def _build_pipeline(
    mock_db,
    mock_recon_repo,
    mock_data_repo,
    mock_config_loader,
    batch_size: int = 100,
) -> IngestionPipeline:
    """Build an IngestionPipeline with mocked dependencies."""
    pipeline = IngestionPipeline(
        db=mock_db,
        config_loader=mock_config_loader,
        batch_size=batch_size,
    )
    pipeline._recon_repo = mock_recon_repo
    pipeline._data_repo = mock_data_repo
    return pipeline


def _rec_date() -> datetime:
    """Return a fixed reconciliation date for tests."""
    return datetime(2024, 1, 15, tzinfo=timezone.utc)


class TestStandardFile:
    """Test with standard fixture containing mix of valid/invalid rows."""

    @pytest.mark.asyncio
    async def test_standard_file_mixed_rows(
        self,
        mock_db: MagicMock,
        mock_reconciliation_file_repo: MagicMock,
        mock_data_container_repo: MagicMock,
        mock_config_loader: MagicMock,
        sample_mapping_config: MappingConfig,
        test_excel_file: str,
    ):
        """Standard file with mixed rows: 8 valid, 2 invalid processed correctly."""
        # Override config to use columns A, B, D, Q with all valid data
        field_mappings = [
            FieldMapping(path="id", column="A", type=FieldMappingType.STRING, required=True),
            FieldMapping(path="trace", column="B", type=FieldMappingType.STRING),
            FieldMapping(path="amount", column="D", type=FieldMappingType.DECIMAL, required=True),
            FieldMapping(
                path="status",
                column="Q",
                type=FieldMappingType.MAPPING,
                mapping={
                    "Thành công": "SUCCESS",
                    "Thất bại": "FAILED",
                    "Đang xử lý": "PENDING",
                    "Đã hoàn tác": "REVERSED",
                },
            ),
            FieldMapping(path="currency", constant="VND", type=FieldMappingType.CONSTANT),
        ]
        mock_config = MappingConfig(
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            sheet_name="Sheet1",
            start_row=2,
            field_mappings=field_mappings,
        )
        mock_config_loader.load_by_partner_type = AsyncMock(return_value=mock_config)

        pipeline = _build_pipeline(
            mock_db, mock_reconciliation_file_repo, mock_data_container_repo, mock_config_loader
        )

        result: IngestionResult = await pipeline.process_file(
            file_path=test_excel_file,
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            reconciliation_date=_rec_date(),
        )

        # 10 rows total: 8 valid, 2 invalid (negative amount, empty ID)
        assert result.stats.total_rows == 10
        assert result.stats.success_rows == 8
        assert result.stats.failed_rows == 2
        assert len(result.errors) >= 2
        assert result.file_record.processing_status == ProcessingStatus.COMPLETED

        # Verify insert_many was called
        assert mock_data_container_repo.insert_many.call_count >= 1


class TestMixedRows:
    """Test 2: Mixed valid/invalid rows."""

    @pytest.mark.asyncio
    async def test_mixed_valid_invalid_rows(
        self,
        mock_db: MagicMock,
        mock_reconciliation_file_repo: MagicMock,
        mock_data_container_repo: MagicMock,
        mock_config_loader: MagicMock,
        sample_mapping_config: MappingConfig,
        test_excel_file: str,
    ):
        """7 valid, 3 invalid rows — stats correct, errors collected."""
        field_mappings = [
            FieldMapping(path="id", column="A", type=FieldMappingType.STRING, required=True),
            FieldMapping(path="trace", column="B", type=FieldMappingType.STRING),
            FieldMapping(path="amount", column="D", type=FieldMappingType.DECIMAL, required=True),
            FieldMapping(
                path="status",
                column="Q",
                type=FieldMappingType.MAPPING,
                mapping={
                    "Thành công": "SUCCESS",
                    "Thất bại": "FAILED",
                    "Đang xử lý": "PENDING",
                    "Đã hoàn tác": "REVERSED",
                },
            ),
            FieldMapping(path="currency", constant="VND", type=FieldMappingType.CONSTANT),
        ]
        mock_config = MappingConfig(
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            sheet_name="Sheet1",
            start_row=2,
            field_mappings=field_mappings,
        )
        mock_config_loader.load_by_partner_type = AsyncMock(return_value=mock_config)

        pipeline = _build_pipeline(
            mock_db, mock_reconciliation_file_repo, mock_data_container_repo, mock_config_loader
        )

        result = await pipeline.process_file(
            file_path=test_excel_file,
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            reconciliation_date=_rec_date(),
        )

        assert result.stats.total_rows == 10
        assert result.stats.success_rows == 8
        assert result.stats.failed_rows == 2
        assert len(result.errors) >= 2

        # Verify error details contain row information
        error_fields = {e.get("field") for e in result.errors}
        assert "amount" in error_fields or "id" in error_fields


class TestAllInvalidRows:
    """Test 3: All invalid rows — processing succeeded, just no valid rows."""

    @pytest.mark.asyncio
    async def test_all_invalid_rows_status_completed(
        self,
        mock_db: MagicMock,
        mock_reconciliation_file_repo: MagicMock,
        mock_data_container_repo: MagicMock,
        mock_config_loader: MagicMock,
        sample_mapping_config: MappingConfig,
        all_invalid_excel_file: str,
    ):
        """0 success, all failed, status COMPLETED (processing succeeded, no valid rows)."""
        field_mappings = [
            FieldMapping(path="id", column="A", type=FieldMappingType.STRING, required=True),
            FieldMapping(path="trace", column="B", type=FieldMappingType.STRING),
            FieldMapping(path="amount", column="D", type=FieldMappingType.DECIMAL, required=True),
            FieldMapping(
                path="status",
                column="Q",
                type=FieldMappingType.MAPPING,
                mapping={
                    "Thành công": "SUCCESS",
                    "Thất bại": "FAILED",
                },
            ),
            FieldMapping(path="currency", constant="VND", type=FieldMappingType.CONSTANT),
        ]
        mock_config = MappingConfig(
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            sheet_name="Sheet1",
            start_row=2,
            field_mappings=field_mappings,
        )
        mock_config_loader.load_by_partner_type = AsyncMock(return_value=mock_config)

        pipeline = _build_pipeline(
            mock_db, mock_reconciliation_file_repo, mock_data_container_repo, mock_config_loader
        )

        result = await pipeline.process_file(
            file_path=all_invalid_excel_file,
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            reconciliation_date=_rec_date(),
        )

        assert result.stats.total_rows == 3
        assert result.stats.success_rows == 0
        assert result.stats.failed_rows == 3
        # Status is COMPLETED because processing itself succeeded
        assert result.file_record.processing_status == ProcessingStatus.COMPLETED
        # No data was inserted
        mock_data_container_repo.insert_many.assert_not_called()


class TestEmptyFile:
    """Test 4: Empty file (header only)."""

    @pytest.mark.asyncio
    async def test_empty_file_header_only(
        self,
        mock_db: MagicMock,
        mock_reconciliation_file_repo: MagicMock,
        mock_data_container_repo: MagicMock,
        mock_config_loader: MagicMock,
        sample_mapping_config: MappingConfig,
        empty_excel_file: str,
    ):
        """Empty file produces total=0, status COMPLETED."""
        field_mappings = [
            FieldMapping(path="id", column="A", type=FieldMappingType.STRING, required=True),
            FieldMapping(path="amount", column="D", type=FieldMappingType.DECIMAL, required=True),
            FieldMapping(path="currency", constant="VND", type=FieldMappingType.CONSTANT),
            FieldMapping(
                path="status",
                column="Q",
                type=FieldMappingType.MAPPING,
                mapping={"Thành công": "SUCCESS"},
            ),
        ]
        mock_config = MappingConfig(
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            sheet_name="Sheet1",
            start_row=2,
            field_mappings=field_mappings,
        )
        mock_config_loader.load_by_partner_type = AsyncMock(return_value=mock_config)

        pipeline = _build_pipeline(
            mock_db, mock_reconciliation_file_repo, mock_data_container_repo, mock_config_loader
        )

        result = await pipeline.process_file(
            file_path=empty_excel_file,
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            reconciliation_date=_rec_date(),
        )

        assert result.stats.total_rows == 0
        assert result.stats.success_rows == 0
        assert result.stats.failed_rows == 0
        assert result.file_record.processing_status == ProcessingStatus.COMPLETED
        assert len(result.errors) == 0
        mock_data_container_repo.insert_many.assert_not_called()


class TestDuplicateDetection:
    """Test 5: Duplicate file hash — early return."""

    @pytest.mark.asyncio
    async def test_duplicate_file_hash_early_return(
        self,
        mock_db: MagicMock,
        mock_reconciliation_file_repo: MagicMock,
        mock_data_container_repo: MagicMock,
        mock_config_loader: MagicMock,
        sample_mapping_config: MappingConfig,
        test_excel_file: str,
    ):
        """Duplicate file hash returns early with error, no processing."""
        # Simulate existing file with same hash
        existing_file = ReconciliationFile(
            partner="MOMO",
            file_name="already_processed.xlsx",
            file_hash="some_hash_value",
            file_type=FileType.SETTLEMENT,
            reconciliation_date=_rec_date(),
        )
        mock_reconciliation_file_repo.find_by_file_hash = AsyncMock(return_value=existing_file)

        pipeline = _build_pipeline(
            mock_db, mock_reconciliation_file_repo, mock_data_container_repo, mock_config_loader
        )

        result = await pipeline.process_file(
            file_path=test_excel_file,
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            reconciliation_date=_rec_date(),
        )

        # Early return — no processing
        assert result.stats.total_rows == 0
        assert result.stats.success_rows == 0
        assert result.stats.failed_rows == 0
        assert len(result.errors) >= 1
        assert "file_duplicate" in result.errors[0].get("field", "")
        # No data inserted
        mock_data_container_repo.insert_many.assert_not_called()
        # No file record created (early return before create)
        mock_reconciliation_file_repo.create.assert_not_called()


class TestBatchInsertion:
    """Test 6: Batch insertion with 250 rows."""

    @pytest.mark.asyncio
    async def test_batch_insertion_250_rows(
        self,
        mock_db: MagicMock,
        mock_reconciliation_file_repo: MagicMock,
        mock_data_container_repo: MagicMock,
        mock_config_loader: MagicMock,
        sample_mapping_config: MappingConfig,
        large_excel_file: str,
    ):
        """250 rows → insert_many called 3 times (100 + 100 + 50)."""
        field_mappings = [
            FieldMapping(path="id", column="A", type=FieldMappingType.STRING, required=True),
            FieldMapping(path="trace", column="B", type=FieldMappingType.STRING),
            FieldMapping(path="amount", column="D", type=FieldMappingType.DECIMAL, required=True),
            FieldMapping(
                path="status",
                column="Q",
                type=FieldMappingType.MAPPING,
                mapping={"Thành công": "SUCCESS"},
            ),
            FieldMapping(path="currency", constant="VND", type=FieldMappingType.CONSTANT),
        ]
        mock_config = MappingConfig(
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            sheet_name="Sheet1",
            start_row=2,
            field_mappings=field_mappings,
        )
        mock_config_loader.load_by_partner_type = AsyncMock(return_value=mock_config)

        pipeline = _build_pipeline(
            mock_db, mock_reconciliation_file_repo, mock_data_container_repo, mock_config_loader,
            batch_size=100,
        )

        result = await pipeline.process_file(
            file_path=large_excel_file,
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            reconciliation_date=_rec_date(),
        )

        assert result.stats.total_rows == 250
        assert result.stats.success_rows == 250

        # insert_many should be called 3 times: 100 + 100 + 50
        assert mock_data_container_repo.insert_many.call_count == 3

        # Verify batch sizes
        call_args_list = mock_data_container_repo.insert_many.call_args_list
        assert len(call_args_list[0][0][0]) == 100
        assert len(call_args_list[1][0][0]) == 100
        assert len(call_args_list[2][0][0]) == 50


class TestExceptionHandling:
    """Test 7: Exception during processing."""

    @pytest.mark.asyncio
    async def test_exception_sets_failed_status(
        self,
        mock_db: MagicMock,
        mock_reconciliation_file_repo: MagicMock,
        mock_data_container_repo: MagicMock,
        mock_config_loader: MagicMock,
        sample_mapping_config: MappingConfig,
        test_excel_file: str,
    ):
        """Exception during processing sets status to FAILED, partial stats returned."""
        # Simulate config loading failure
        mock_config_loader.load_by_partner_type = AsyncMock(
            side_effect=Exception("Config load failed")
        )

        pipeline = _build_pipeline(
            mock_db, mock_reconciliation_file_repo, mock_data_container_repo, mock_config_loader
        )

        result = await pipeline.process_file(
            file_path=test_excel_file,
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            reconciliation_date=_rec_date(),
        )

        assert result.file_record.processing_status == ProcessingStatus.FAILED
        assert len(result.errors) >= 1
        assert any("Config load failed" in e.get("reason", "") for e in result.errors)

        # Verify status update was attempted
        mock_reconciliation_file_repo.update_status.assert_called()
        status_call = mock_reconciliation_file_repo.update_status.call_args
        assert status_call[0][1] == ProcessingStatus.FAILED


class TestSourceFileIdLinking:
    """Test 8: DataContainer records have correct source_file_id."""

    @pytest.mark.asyncio
    async def test_data_container_source_file_id_linking(
        self,
        mock_db: MagicMock,
        mock_reconciliation_file_repo: MagicMock,
        mock_data_container_repo: MagicMock,
        mock_config_loader: MagicMock,
        sample_mapping_config: MappingConfig,
    ):
        """DataContainer records link to the correct ReconciliationFile via source_file_id."""
        import openpyxl

        # Column indices (1-based): A=1, B=2, D=4, Q=17
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            # Row 1: header
            ws.cell(row=1, column=1, value="TransactionID")
            ws.cell(row=1, column=2, value="Trace")
            ws.cell(row=1, column=4, value="Amount")
            ws.cell(row=1, column=17, value="Status")
            # Rows 2-4: data
            ws.cell(row=2, column=1, value="TXN001")
            ws.cell(row=2, column=2, value="TRACE001")
            ws.cell(row=2, column=4, value=100000)
            ws.cell(row=2, column=17, value="Thành công")
            ws.cell(row=3, column=1, value="TXN002")
            ws.cell(row=3, column=2, value="TRACE002")
            ws.cell(row=3, column=4, value=200000)
            ws.cell(row=3, column=17, value="Thành công")
            ws.cell(row=4, column=1, value="TXN003")
            ws.cell(row=4, column=2, value="TRACE003")
            ws.cell(row=4, column=4, value=300000)
            ws.cell(row=4, column=17, value="Thành công")
            wb.save(f.name)
            temp_path = f.name

        try:
            field_mappings = [
                FieldMapping(path="id", column="A", type=FieldMappingType.STRING, required=True),
                FieldMapping(path="trace", column="B", type=FieldMappingType.STRING),
                FieldMapping(path="amount", column="D", type=FieldMappingType.DECIMAL, required=True),
                FieldMapping(
                    path="status",
                    column="Q",
                    type=FieldMappingType.MAPPING,
                    mapping={"Thành công": "SUCCESS"},
                ),
                FieldMapping(path="currency", constant="VND", type=FieldMappingType.CONSTANT),
            ]
            mock_config = MappingConfig(
                partner="MOMO",
                workflow_type="UPC",
                file_type=FileType.SETTLEMENT,
                sheet_name="Sheet1",
                start_row=2,
                field_mappings=field_mappings,
            )
            mock_config_loader.load_by_partner_type = AsyncMock(return_value=mock_config)

            pipeline = _build_pipeline(
                mock_db, mock_reconciliation_file_repo, mock_data_container_repo, mock_config_loader
            )

            result = await pipeline.process_file(
                file_path=temp_path,
                partner="MOMO",
                workflow_type="UPC",
                file_type=FileType.SETTLEMENT,
                reconciliation_date=_rec_date(),
            )

            assert result.stats.success_rows == 3

            # Get the file_record ID that was created
            file_record_id = result.file_record.id
            assert file_record_id is not None

            # Verify insert_many was called with DataContainers linked to this file
            assert mock_data_container_repo.insert_many.call_count == 1
            batch = mock_data_container_repo.insert_many.call_args[0][0]
            assert len(batch) == 3

            for container in batch:
                assert isinstance(container, DataContainer)
                assert container.source_file_id == file_record_id
        finally:
            Path(temp_path).unlink(missing_ok=True)


class TestPartnerDataNesting:
    """Test 9: Partner data correctly nested in DataContainer."""

    @pytest.mark.asyncio
    async def test_partner_data_nesting(
        self,
        mock_db: MagicMock,
        mock_reconciliation_file_repo: MagicMock,
        mock_data_container_repo: MagicMock,
        mock_config_loader: MagicMock,
        sample_mapping_config: MappingConfig,
    ):
        """Partner data (amount, currency, status, trace) correctly nested in DataContainer."""
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            # Row 1: header
            ws.cell(row=1, column=1, value="TransactionID")
            ws.cell(row=1, column=2, value="Trace")
            ws.cell(row=1, column=4, value="Amount")
            ws.cell(row=1, column=17, value="Status")
            # Row 2: valid — Thành công
            ws.cell(row=2, column=1, value="TXN001")
            ws.cell(row=2, column=2, value="TRACE001")
            ws.cell(row=2, column=4, value=150000)
            ws.cell(row=2, column=17, value="Thành công")
            # Row 3: valid — Thất bại
            ws.cell(row=3, column=1, value="TXN002")
            ws.cell(row=3, column=2, value="TRACE002")
            ws.cell(row=3, column=4, value=250000)
            ws.cell(row=3, column=17, value="Thất bại")
            wb.save(f.name)
            temp_path = f.name

        try:
            field_mappings = [
                FieldMapping(path="id", column="A", type=FieldMappingType.STRING, required=True),
                FieldMapping(path="trace", column="B", type=FieldMappingType.STRING),
                FieldMapping(path="amount", column="D", type=FieldMappingType.DECIMAL, required=True),
                FieldMapping(
                    path="status",
                    column="Q",
                    type=FieldMappingType.MAPPING,
                    mapping={
                        "Thành công": "SUCCESS",
                        "Thất bại": "FAILED",
                    },
                ),
                FieldMapping(path="currency", constant="VND", type=FieldMappingType.CONSTANT),
            ]
            mock_config = MappingConfig(
                partner="MOMO",
                workflow_type="UPC",
                file_type=FileType.SETTLEMENT,
                sheet_name="Sheet1",
                start_row=2,
                field_mappings=field_mappings,
            )
            mock_config_loader.load_by_partner_type = AsyncMock(return_value=mock_config)

            pipeline = _build_pipeline(
                mock_db, mock_reconciliation_file_repo, mock_data_container_repo, mock_config_loader
            )

            result = await pipeline.process_file(
                file_path=temp_path,
                partner="MOMO",
                workflow_type="UPC",
                file_type=FileType.SETTLEMENT,
                reconciliation_date=_rec_date(),
            )

            assert result.stats.success_rows == 2

            batch = mock_data_container_repo.insert_many.call_args[0][0]
            assert len(batch) == 2

            # Check first container
            container1 = batch[0]
            assert isinstance(container1.partner_data, PartnerData)
            assert container1.partner_data.trace == "TRACE001"
            assert container1.partner_data.amount == Decimal("150000")
            assert container1.partner_data.currency == "VND"
            assert container1.partner_data.status == "SUCCESS"

            # Check second container
            container2 = batch[1]
            assert isinstance(container2.partner_data, PartnerData)
            assert container2.partner_data.trace == "TRACE002"
            assert container2.partner_data.amount == Decimal("250000")
            assert container2.partner_data.currency == "VND"
            assert container2.partner_data.status == "FAILED"
        finally:
            Path(temp_path).unlink(missing_ok=True)


class TestStatsAccuracy:
    """Test 10: ProcessingStats accuracy."""

    @pytest.mark.asyncio
    async def test_processing_stats_accuracy(
        self,
        mock_db: MagicMock,
        mock_reconciliation_file_repo: MagicMock,
        mock_data_container_repo: MagicMock,
        mock_config_loader: MagicMock,
        sample_mapping_config: MappingConfig,
        test_excel_file: str,
    ):
        """ProcessingStats returned matches ReconciliationFile stats."""
        field_mappings = [
            FieldMapping(path="id", column="A", type=FieldMappingType.STRING, required=True),
            FieldMapping(path="trace", column="B", type=FieldMappingType.STRING),
            FieldMapping(path="amount", column="D", type=FieldMappingType.DECIMAL, required=True),
            FieldMapping(
                path="status",
                column="Q",
                type=FieldMappingType.MAPPING,
                mapping={
                    "Thành công": "SUCCESS",
                    "Thất bại": "FAILED",
                    "Đang xử lý": "PENDING",
                    "Đã hoàn tác": "REVERSED",
                },
            ),
            FieldMapping(path="currency", constant="VND", type=FieldMappingType.CONSTANT),
        ]
        mock_config = MappingConfig(
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            sheet_name="Sheet1",
            start_row=2,
            field_mappings=field_mappings,
        )
        mock_config_loader.load_by_partner_type = AsyncMock(return_value=mock_config)

        pipeline = _build_pipeline(
            mock_db, mock_reconciliation_file_repo, mock_data_container_repo, mock_config_loader
        )

        result = await pipeline.process_file(
            file_path=test_excel_file,
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            reconciliation_date=_rec_date(),
        )

        # IngestionResult.stats match expected
        assert isinstance(result.stats, ProcessingStats)
        assert result.stats.total_rows == 10
        assert result.stats.success_rows == 8
        assert result.stats.failed_rows == 2

        # ReconciliationFile stats match ProcessingStats
        file_record = result.file_record
        assert file_record.total_rows == result.stats.total_rows
        assert file_record.success_rows == result.stats.success_rows
        assert file_record.failed_rows == result.stats.failed_rows

        # Verify update_processing_stats was called with matching values
        mock_reconciliation_file_repo.update_processing_stats.assert_called_once()
        stats_call = mock_reconciliation_file_repo.update_processing_stats.call_args
        assert stats_call[0][1] == result.stats.total_rows
        assert stats_call[0][2] == result.stats.success_rows
        assert stats_call[0][3] == result.stats.failed_rows
