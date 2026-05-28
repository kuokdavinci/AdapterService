"""Tests for IngestionPipeline — TDD red/green/refactor cycle.

Tests cover:
- IngestionResult dataclass construction
- _tuple_to_dict conversion
- _compute_file_hash returns consistent SHA256
- process_file happy path (all rows valid)
- process_file with mixed valid/invalid rows
- process_file with duplicate file hash (early return)
- process_file exception handling (status → FAILED)
- Batch insertion (verify insert_many called with correct batch)
- StructuredLogger integration (lifecycle events emitted)
"""

import hashlib
import uuid
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from src.core.enums import FileType, ProcessingStatus, TransactionStatus
from src.core.types import (
    CanonicalTransaction,
    FieldMapping,
    FieldMappingType,
    ProcessingStats,
    ValidationError,
)


class MockStructuredLogger:
    """Mock logger that captures emitted events for test verification."""

    def __init__(self):
        self.events = []

    def emit_file_started(self, file_id, file_name, partner):
        self.events.append(("FILE_STARTED", {"file_id": file_id, "file_name": file_name, "partner": partner}))

    def emit_file_completed(self, file_id, total, success, failed, duration_ms):
        self.events.append(("FILE_COMPLETED", {"file_id": file_id, "total": total, "success": success, "failed": failed, "duration_ms": duration_ms}))

    def emit_file_failed(self, file_id, error):
        self.events.append(("FILE_FAILED", {"file_id": file_id, "error": error}))

    def emit_row_success(self, file_id, row_number, trace):
        self.events.append(("ROW_SUCCESS", {"file_id": file_id, "row_number": row_number, "trace": trace}))

    def emit_row_failed(self, file_id, row_number, trace, reason):
        self.events.append(("ROW_FAILED", {"file_id": file_id, "row_number": row_number, "trace": trace, "reason": reason}))


class TestIngestionResult:
    """Tests for IngestionResult dataclass."""

    def test_ingestion_result_construction(self):
        """IngestionResult can be constructed with file_record, stats, errors."""
        from src.pipeline import IngestionResult
        from src.models.reconciliation_file import ReconciliationFile

        now = datetime.now(timezone.utc)
        file_record = ReconciliationFile(
            partner="MOMO",
            file_name="test.xlsx",
            file_hash="abc123",
            file_type=FileType.SETTLEMENT,
            reconciliation_date=now,
        )
        stats = ProcessingStats(total_rows=10, success_rows=8, failed_rows=2)
        errors = [{"row": 3, "field": "amount", "reason": "invalid"}]

        result = IngestionResult(
            file_record=file_record, stats=stats, errors=errors
        )

        assert result.file_record is file_record
        assert result.stats == stats
        assert result.errors == errors


class TestTupleToDict:
    """Tests for _tuple_to_dict conversion."""

    def test_tuple_to_dict_basic(self):
        """Tuple index 0 → 'A', 1 → 'B', etc."""
        from src.pipeline import IngestionPipeline

        pipeline = IngestionPipeline(db=MagicMock(), config_loader=MagicMock())
        row_tuple = ("TXN001", "100.50", "VND", "SUCCESS")
        result = pipeline._tuple_to_dict(row_tuple)

        assert result == {
            "A": "TXN001",
            "B": "100.50",
            "C": "VND",
            "D": "SUCCESS",
        }

    def test_tuple_to_dict_empty(self):
        """Empty tuple produces empty dict."""
        from src.pipeline import IngestionPipeline

        pipeline = IngestionPipeline(db=MagicMock(), config_loader=MagicMock())
        result = pipeline._tuple_to_dict(())
        assert result == {}

    def test_tuple_to_dict_single_element(self):
        """Single element tuple maps to 'A'."""
        from src.pipeline import IngestionPipeline

        pipeline = IngestionPipeline(db=MagicMock(), config_loader=MagicMock())
        result = pipeline._tuple_to_dict(("only_value",))
        assert result == {"A": "only_value"}


class TestComputeFileHash:
    """Tests for _compute_file_hash."""

    @pytest.mark.asyncio
    async def test_compute_file_hash_consistent(self, tmp_path):
        """_compute_file_hash returns consistent SHA256 for same content."""
        from src.pipeline import IngestionPipeline

        file_path = tmp_path / "file.xlsx"
        file_path.write_bytes(b"test content")
        pipeline = IngestionPipeline(db=MagicMock(), config_loader=MagicMock())
        hash1 = await pipeline._compute_file_hash(str(file_path))
        hash2 = await pipeline._compute_file_hash(str(file_path))
        assert hash1 == hash2

    @pytest.mark.asyncio
    async def test_compute_file_hash_different_content(self, tmp_path):
        """Different file content produces different hashes."""
        from src.pipeline import IngestionPipeline

        file_a = tmp_path / "a.xlsx"
        file_b = tmp_path / "b.xlsx"
        file_a.write_bytes(b"content a")
        file_b.write_bytes(b"content b")
        pipeline = IngestionPipeline(db=MagicMock(), config_loader=MagicMock())
        hash1 = await pipeline._compute_file_hash(str(file_a))
        hash2 = await pipeline._compute_file_hash(str(file_b))
        assert hash1 != hash2

    @pytest.mark.asyncio
    async def test_compute_file_hash_is_sha256(self, tmp_path):
        """_compute_file_hash produces a valid SHA256 hex string."""
        from src.pipeline import IngestionPipeline

        file_path = tmp_path / "file.xlsx"
        file_path.write_bytes(b"test")
        pipeline = IngestionPipeline(db=MagicMock(), config_loader=MagicMock())
        result = await pipeline._compute_file_hash(str(file_path))
        assert len(result) == 64
        assert all(c in "0123456789abcdef" for c in result)


class TestProcessFileHappyPath:
    """Tests for process_file happy path — all rows valid."""

    @pytest.mark.asyncio
    async def test_process_file_all_rows_valid(self):
        """process_file processes all rows successfully with correct stats."""
        from src.pipeline import IngestionPipeline, IngestionResult
        from src.models.reconciliation_file import (
            ReconciliationFile,
            ReconciliationFileRepository,
        )
        from src.models.data_container import DataContainerRepository
        from src.models.mapping_config import MappingConfig
        from src.config.loader import ConfigLoader

        # Build a mock config with field mappings
        field_mappings = [
            FieldMapping(path="id", column="A", type=FieldMappingType.STRING, required=True),
            FieldMapping(path="amount", column="B", type=FieldMappingType.DECIMAL, required=True),
            FieldMapping(path="currency", constant="VND", type=FieldMappingType.CONSTANT),
            FieldMapping(
                path="status",
                column="C",
                type=FieldMappingType.MAPPING,
                mapping={"Success": "SUCCESS", "Failed": "FAILED"},
            ),
        ]

        mock_config = MappingConfig(
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            sheet_name="Sheet1",
            start_row=2,
            field_mappings=field_mappings,
        )

        # Mock config loader
        mock_config_loader = MagicMock(spec=ConfigLoader)
        mock_config_loader.load_by_partner_type = AsyncMock(return_value=mock_config)

        # Mock repositories
        mock_db = MagicMock()
        mock_recon_repo = MagicMock(spec=ReconciliationFileRepository)
        mock_recon_repo.find_by_file_hash = AsyncMock(return_value=None)
        mock_recon_repo.create = AsyncMock(side_effect=lambda doc: doc)
        mock_recon_repo.update_processing_stats = AsyncMock(return_value=True)
        mock_recon_repo.update_status = AsyncMock(return_value=True)

        mock_data_repo = MagicMock(spec=DataContainerRepository)
        mock_data_repo.insert_many = AsyncMock(return_value=3)

        # Wire up db to return our mock repos
        mock_db.__getitem__ = MagicMock(side_effect=lambda name: MagicMock())

        pipeline = IngestionPipeline(
            db=mock_db, config_loader=mock_config_loader, batch_size=100
        )
        # Override repos with our mocks
        pipeline._recon_repo = mock_recon_repo
        pipeline._data_repo = mock_data_repo

        # Create a temporary Excel file for testing
        import tempfile
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            # Row 1: header (skipped by start_row=2)
            ws.append(["ID", "Amount", "Status"])
            # Rows 2-4: data
            ws.append(["TXN001", 100, "Success"])
            ws.append(["TXN002", 200, "Success"])
            ws.append(["TXN003", 300, "Success"])
            wb.save(f.name)
            temp_path = f.name

        try:
            rec_date = datetime(2024, 1, 15, tzinfo=timezone.utc)
            result: IngestionResult = await pipeline.process_file(
                file_path=temp_path,
                partner="MOMO",
                workflow_type="UPC",
                file_type=FileType.SETTLEMENT,
                reconciliation_date=rec_date,
            )

            assert result is not None
            assert result.stats.total_rows == 3
            assert result.stats.success_rows == 3
            assert result.stats.failed_rows == 0
            assert len(result.errors) == 0
            assert result.file_record.processing_status == ProcessingStatus.COMPLETED

            # Verify batch insertion was called
            mock_data_repo.insert_many.assert_called_once()
            batch = mock_data_repo.insert_many.call_args[0][0]
            assert len(batch) == 3
        finally:
            Path(temp_path).unlink(missing_ok=True)


class TestProcessFileMixedRows:
    """Tests for process_file with mixed valid/invalid rows."""

    @pytest.mark.asyncio
    async def test_process_file_mixed_valid_invalid_rows(self):
        """Invalid rows are collected as errors without stopping the pipeline."""
        from src.pipeline import IngestionPipeline
        from src.models.reconciliation_file import ReconciliationFileRepository
        from src.models.data_container import DataContainerRepository
        from src.models.mapping_config import MappingConfig
        from src.config.loader import ConfigLoader

        field_mappings = [
            FieldMapping(path="id", column="A", type=FieldMappingType.STRING, required=True),
            FieldMapping(path="amount", column="B", type=FieldMappingType.DECIMAL, required=True),
            FieldMapping(path="currency", constant="VND", type=FieldMappingType.CONSTANT),
            FieldMapping(
                path="status",
                column="C",
                type=FieldMappingType.MAPPING,
                mapping={"Success": "SUCCESS", "Failed": "FAILED"},
            ),
        ]

        mock_config = MappingConfig(
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            sheet_name="Sheet1",
            start_row=2,
            field_mappings=field_mappings,
        )

        mock_config_loader = MagicMock(spec=ConfigLoader)
        mock_config_loader.load_by_partner_type = AsyncMock(return_value=mock_config)

        mock_db = MagicMock()
        mock_recon_repo = MagicMock(spec=ReconciliationFileRepository)
        mock_recon_repo.find_by_file_hash = AsyncMock(return_value=None)
        mock_recon_repo.create = AsyncMock(side_effect=lambda doc: doc)
        mock_recon_repo.update_processing_stats = AsyncMock(return_value=True)
        mock_recon_repo.update_status = AsyncMock(return_value=True)

        mock_data_repo = MagicMock(spec=DataContainerRepository)
        mock_data_repo.insert_many = AsyncMock(return_value=2)

        mock_db.__getitem__ = MagicMock(side_effect=lambda name: MagicMock())

        pipeline = IngestionPipeline(
            db=mock_db, config_loader=mock_config_loader, batch_size=100
        )
        pipeline._recon_repo = mock_recon_repo
        pipeline._data_repo = mock_data_repo

        import tempfile
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            # Row 1: header (skipped by start_row=2)
            ws.append(["ID", "Amount", "Status"])
            # Row 2: valid
            ws.append(["TXN001", 100, "Success"])
            # Row 3: invalid amount (empty)
            ws.append(["TXN002", None, "Success"])
            # Row 4: valid
            ws.append(["TXN003", 300, "Success"])
            wb.save(f.name)
            temp_path = f.name

        try:
            rec_date = datetime(2024, 1, 15, tzinfo=timezone.utc)
            result = await pipeline.process_file(
                file_path=temp_path,
                partner="MOMO",
                workflow_type="UPC",
                file_type=FileType.SETTLEMENT,
                reconciliation_date=rec_date,
            )

            assert result.stats.total_rows == 3
            assert result.stats.success_rows == 2
            assert result.stats.failed_rows == 1
            assert len(result.errors) >= 1
        finally:
            Path(temp_path).unlink(missing_ok=True)


class TestProcessFileDuplicate:
    """Tests for process_file with duplicate file hash."""

    @pytest.mark.asyncio
    async def test_process_file_duplicate_hash_early_return(self):
        """Duplicate file hash returns early with error."""
        from src.pipeline import IngestionPipeline
        from src.models.reconciliation_file import ReconciliationFile, ReconciliationFileRepository
        from src.config.loader import ConfigLoader

        mock_db = MagicMock()
        mock_recon_repo = MagicMock(spec=ReconciliationFileRepository)

        existing_file = ReconciliationFile(
            partner="MOMO",
            file_name="duplicate.xlsx",
            file_hash="some_hash",
            file_type=FileType.SETTLEMENT,
            reconciliation_date=datetime(2024, 1, 15, tzinfo=timezone.utc),
        )
        mock_recon_repo.find_by_file_hash = AsyncMock(return_value=existing_file)

        mock_config_loader = MagicMock(spec=ConfigLoader)

        pipeline = IngestionPipeline(
            db=mock_db, config_loader=mock_config_loader, batch_size=100
        )
        pipeline._recon_repo = mock_recon_repo

        rec_date = datetime(2024, 1, 15, tzinfo=timezone.utc)
        result = await pipeline.process_file(
            file_path="/some/path/file.xlsx",
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            reconciliation_date=rec_date,
        )

        # Should return early with error
        assert result.stats.total_rows == 0
        assert result.stats.success_rows == 0
        assert result.stats.failed_rows == 0
        assert len(result.errors) >= 1


class TestProcessFileException:
    """Tests for process_file exception handling."""

    @pytest.mark.asyncio
    async def test_process_file_exception_sets_failed_status(self, tmp_path):
        """Exception during processing sets status to FAILED."""
        from src.pipeline import IngestionPipeline
        from src.models.reconciliation_file import ReconciliationFileRepository
        from src.config.loader import ConfigLoader

        # Create a real temp file so _compute_file_hash succeeds
        excel_file = tmp_path / "file.xlsx"
        excel_file.write_bytes(b"fake excel content")

        mock_db = MagicMock()
        mock_recon_repo = MagicMock(spec=ReconciliationFileRepository)
        mock_recon_repo.find_by_file_hash = AsyncMock(return_value=None)
        mock_recon_repo.create = AsyncMock(side_effect=lambda doc: doc)
        mock_recon_repo.update_processing_stats = AsyncMock(return_value=True)
        mock_recon_repo.update_status = AsyncMock(return_value=True)

        mock_config_loader = MagicMock(spec=ConfigLoader)
        # Simulate config loading failure
        mock_config_loader.load_by_partner_type = AsyncMock(
            side_effect=Exception("Config load failed")
        )

        pipeline = IngestionPipeline(
            db=mock_db, config_loader=mock_config_loader, batch_size=100
        )
        pipeline._recon_repo = mock_recon_repo

        rec_date = datetime(2024, 1, 15, tzinfo=timezone.utc)
        result = await pipeline.process_file(
            file_path=str(excel_file),
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            reconciliation_date=rec_date,
        )

        assert result.file_record is not None
        assert result.file_record.processing_status == ProcessingStatus.FAILED


class TestBatchInsertion:
    """Tests for batch insertion behavior."""

    @pytest.mark.asyncio
    async def test_batch_insertion_called_with_correct_batch_size(self):
        """insert_many is called when batch buffer reaches batch_size."""
        from src.pipeline import IngestionPipeline
        from src.models.reconciliation_file import ReconciliationFileRepository
        from src.models.data_container import DataContainerRepository
        from src.models.mapping_config import MappingConfig
        from src.config.loader import ConfigLoader

        field_mappings = [
            FieldMapping(path="id", column="A", type=FieldMappingType.STRING, required=True),
            FieldMapping(path="amount", column="B", type=FieldMappingType.DECIMAL, required=True),
            FieldMapping(path="currency", constant="VND", type=FieldMappingType.CONSTANT),
            FieldMapping(
                path="status",
                column="C",
                type=FieldMappingType.MAPPING,
                mapping={"Success": "SUCCESS"},
            ),
        ]

        mock_config = MappingConfig(
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            sheet_name="Sheet1",
            start_row=2,
            field_mappings=field_mappings,
        )

        mock_config_loader = MagicMock(spec=ConfigLoader)
        mock_config_loader.load_by_partner_type = AsyncMock(return_value=mock_config)

        mock_db = MagicMock()
        mock_recon_repo = MagicMock(spec=ReconciliationFileRepository)
        mock_recon_repo.find_by_file_hash = AsyncMock(return_value=None)
        mock_recon_repo.create = AsyncMock(side_effect=lambda doc: doc)
        mock_recon_repo.update_processing_stats = AsyncMock(return_value=True)
        mock_recon_repo.update_status = AsyncMock(return_value=True)

        mock_data_repo = MagicMock(spec=DataContainerRepository)
        mock_data_repo.insert_many = AsyncMock(side_effect=lambda batch: len(batch))

        mock_db.__getitem__ = MagicMock(side_effect=lambda name: MagicMock())

        # Use batch_size=5 to test batching
        pipeline = IngestionPipeline(
            db=mock_db, config_loader=mock_config_loader, batch_size=5
        )
        pipeline._recon_repo = mock_recon_repo
        pipeline._data_repo = mock_data_repo

        import tempfile
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            # Row 1: header (skipped by start_row=2)
            ws.append(["ID", "Amount", "Status"])
            # 7 rows — should trigger one batch of 5, then a final flush of 2
            for i in range(7):
                ws.append([f"TXN{i:03d}", 100 + i, "Success"])
            wb.save(f.name)
            temp_path = f.name

        try:
            rec_date = datetime(2024, 1, 15, tzinfo=timezone.utc)
            result = await pipeline.process_file(
                file_path=temp_path,
                partner="MOMO",
                workflow_type="UPC",
                file_type=FileType.SETTLEMENT,
                reconciliation_date=rec_date,
            )

            assert result.stats.total_rows == 7
            assert result.stats.success_rows == 7

            # Should be called twice: once for batch of 5, once for remaining 2
            assert mock_data_repo.insert_many.call_count == 2
            first_batch = mock_data_repo.insert_many.call_args_list[0][0][0]
            second_batch = mock_data_repo.insert_many.call_args_list[1][0][0]
            assert len(first_batch) == 5
            assert len(second_batch) == 2
        finally:
            Path(temp_path).unlink(missing_ok=True)


class TestPipelineLogging:
    """Tests for StructuredLogger integration in IngestionPipeline."""

    @pytest.mark.asyncio
    async def test_logger_emits_events_happy_path(self):
        """process_file emits FILE_STARTED → ROW_SUCCESS×N → FILE_COMPLETED."""
        from src.pipeline import IngestionPipeline
        from src.models.reconciliation_file import ReconciliationFileRepository
        from src.models.data_container import DataContainerRepository
        from src.models.mapping_config import MappingConfig
        from src.config.loader import ConfigLoader

        field_mappings = [
            FieldMapping(path="id", column="A", type=FieldMappingType.STRING, required=True),
            FieldMapping(path="amount", column="B", type=FieldMappingType.DECIMAL, required=True),
            FieldMapping(path="currency", constant="VND", type=FieldMappingType.CONSTANT),
            FieldMapping(
                path="status",
                column="C",
                type=FieldMappingType.MAPPING,
                mapping={"Success": "SUCCESS"},
            ),
        ]

        mock_config = MappingConfig(
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            sheet_name="Sheet1",
            start_row=2,
            field_mappings=field_mappings,
        )

        mock_config_loader = MagicMock(spec=ConfigLoader)
        mock_config_loader.load_by_partner_type = AsyncMock(return_value=mock_config)

        mock_db = MagicMock()
        mock_recon_repo = MagicMock(spec=ReconciliationFileRepository)
        mock_recon_repo.find_by_file_hash = AsyncMock(return_value=None)
        mock_recon_repo.create = AsyncMock(side_effect=lambda doc: doc)
        mock_recon_repo.update_processing_stats = AsyncMock(return_value=True)
        mock_recon_repo.update_status = AsyncMock(return_value=True)

        mock_data_repo = MagicMock(spec=DataContainerRepository)
        mock_data_repo.insert_many = AsyncMock(return_value=3)

        mock_db.__getitem__ = MagicMock(side_effect=lambda name: MagicMock())

        mock_logger = MockStructuredLogger()

        pipeline = IngestionPipeline(
            db=mock_db, config_loader=mock_config_loader, batch_size=100, logger=mock_logger,
        )
        pipeline._recon_repo = mock_recon_repo
        pipeline._data_repo = mock_data_repo

        import tempfile
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(["ID", "Amount", "Status"])
            ws.append(["TXN001", 100, "Success"])
            ws.append(["TXN002", 200, "Success"])
            ws.append(["TXN003", 300, "Success"])
            wb.save(f.name)
            temp_path = f.name

        try:
            rec_date = datetime(2024, 1, 15, tzinfo=timezone.utc)
            result = await pipeline.process_file(
                file_path=temp_path,
                partner="MOMO",
                workflow_type="UPC",
                file_type=FileType.SETTLEMENT,
                reconciliation_date=rec_date,
            )

            assert result.stats.total_rows == 3
            assert result.stats.success_rows == 3

            events = mock_logger.events
            assert events[0][0] == "FILE_STARTED"
            row_success_events = [e for e in events if e[0] == "ROW_SUCCESS"]
            assert len(row_success_events) == 3
            assert events[-1][0] == "FILE_COMPLETED"
            completed_data = events[-1][1]
            assert completed_data["total"] == 3
            assert completed_data["success"] == 3
            assert completed_data["failed"] == 0
            assert completed_data["duration_ms"] > 0
        finally:
            Path(temp_path).unlink(missing_ok=True)

    @pytest.mark.asyncio
    async def test_logger_emits_mixed_rows(self):
        """process_file emits ROW_SUCCESS and ROW_FAILED for mixed rows."""
        from src.pipeline import IngestionPipeline
        from src.models.reconciliation_file import ReconciliationFileRepository
        from src.models.data_container import DataContainerRepository
        from src.models.mapping_config import MappingConfig
        from src.config.loader import ConfigLoader

        field_mappings = [
            FieldMapping(path="id", column="A", type=FieldMappingType.STRING, required=True),
            FieldMapping(path="amount", column="B", type=FieldMappingType.DECIMAL, required=True),
            FieldMapping(path="currency", constant="VND", type=FieldMappingType.CONSTANT),
            FieldMapping(
                path="status",
                column="C",
                type=FieldMappingType.MAPPING,
                mapping={"Success": "SUCCESS", "Failed": "FAILED"},
            ),
        ]

        mock_config = MappingConfig(
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            sheet_name="Sheet1",
            start_row=2,
            field_mappings=field_mappings,
        )

        mock_config_loader = MagicMock(spec=ConfigLoader)
        mock_config_loader.load_by_partner_type = AsyncMock(return_value=mock_config)

        mock_db = MagicMock()
        mock_recon_repo = MagicMock(spec=ReconciliationFileRepository)
        mock_recon_repo.find_by_file_hash = AsyncMock(return_value=None)
        mock_recon_repo.create = AsyncMock(side_effect=lambda doc: doc)
        mock_recon_repo.update_processing_stats = AsyncMock(return_value=True)
        mock_recon_repo.update_status = AsyncMock(return_value=True)

        mock_data_repo = MagicMock(spec=DataContainerRepository)
        mock_data_repo.insert_many = AsyncMock(return_value=2)

        mock_db.__getitem__ = MagicMock(side_effect=lambda name: MagicMock())

        mock_logger = MockStructuredLogger()

        pipeline = IngestionPipeline(
            db=mock_db, config_loader=mock_config_loader, batch_size=100, logger=mock_logger,
        )
        pipeline._recon_repo = mock_recon_repo
        pipeline._data_repo = mock_data_repo

        import tempfile
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(["ID", "Amount", "Status"])
            ws.append(["TXN001", 100, "Success"])
            ws.append(["TXN002", None, "Success"])  # invalid amount
            ws.append(["TXN003", 300, "Success"])
            wb.save(f.name)
            temp_path = f.name

        try:
            rec_date = datetime(2024, 1, 15, tzinfo=timezone.utc)
            result = await pipeline.process_file(
                file_path=temp_path,
                partner="MOMO",
                workflow_type="UPC",
                file_type=FileType.SETTLEMENT,
                reconciliation_date=rec_date,
            )

            assert result.stats.total_rows == 3
            assert result.stats.success_rows == 2
            assert result.stats.failed_rows == 1

            events = mock_logger.events
            assert events[0][0] == "FILE_STARTED"
            row_success = [e for e in events if e[0] == "ROW_SUCCESS"]
            row_failed = [e for e in events if e[0] == "ROW_FAILED"]
            assert len(row_success) == 2
            assert len(row_failed) == 1
            assert row_failed[0][1]["reason"]  # reason is non-empty
            assert events[-1][0] == "FILE_COMPLETED"
            completed_data = events[-1][1]
            assert completed_data["total"] == 3
            assert completed_data["success"] == 2
            assert completed_data["failed"] == 1
        finally:
            Path(temp_path).unlink(missing_ok=True)

    @pytest.mark.asyncio
    async def test_logger_duplicate_emits_file_failed_only(self, tmp_path):
        """Duplicate file emits FILE_FAILED without FILE_STARTED."""
        from src.pipeline import IngestionPipeline
        from src.models.reconciliation_file import ReconciliationFile, ReconciliationFileRepository
        from src.config.loader import ConfigLoader

        # Create a real file so _compute_file_hash succeeds
        excel_file = tmp_path / "duplicate.xlsx"
        excel_file.write_bytes(b"fake excel content")

        mock_db = MagicMock()
        mock_recon_repo = MagicMock(spec=ReconciliationFileRepository)

        existing_file = ReconciliationFile(
            partner="MOMO",
            file_name="duplicate.xlsx",
            file_hash="some_hash",
            file_type=FileType.SETTLEMENT,
            reconciliation_date=datetime(2024, 1, 15, tzinfo=timezone.utc),
        )
        mock_recon_repo.find_by_file_hash = AsyncMock(return_value=existing_file)

        mock_config_loader = MagicMock(spec=ConfigLoader)

        mock_logger = MockStructuredLogger()

        pipeline = IngestionPipeline(
            db=mock_db, config_loader=mock_config_loader, batch_size=100, logger=mock_logger,
        )
        pipeline._recon_repo = mock_recon_repo

        rec_date = datetime(2024, 1, 15, tzinfo=timezone.utc)
        result = await pipeline.process_file(
            file_path=str(excel_file),
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            reconciliation_date=rec_date,
        )

        assert result.stats.total_rows == 0
        events = mock_logger.events
        assert len(events) == 1
        assert events[0][0] == "FILE_FAILED"
        assert "already processed" in events[0][1]["error"].lower()
        assert events[0][1]["file_id"] == "duplicate"

    @pytest.mark.asyncio
    async def test_logger_exception_emits_file_failed(self, tmp_path):
        """Exception during processing emits FILE_FAILED with error message."""
        from src.pipeline import IngestionPipeline
        from src.models.reconciliation_file import ReconciliationFileRepository
        from src.config.loader import ConfigLoader

        excel_file = tmp_path / "file.xlsx"
        excel_file.write_bytes(b"fake excel content")

        mock_db = MagicMock()
        mock_recon_repo = MagicMock(spec=ReconciliationFileRepository)
        mock_recon_repo.find_by_file_hash = AsyncMock(return_value=None)
        mock_recon_repo.create = AsyncMock(side_effect=lambda doc: doc)
        mock_recon_repo.update_processing_stats = AsyncMock(return_value=True)
        mock_recon_repo.update_status = AsyncMock(return_value=True)

        mock_config_loader = MagicMock(spec=ConfigLoader)
        mock_config_loader.load_by_partner_type = AsyncMock(
            side_effect=Exception("Config load failed")
        )

        mock_logger = MockStructuredLogger()

        pipeline = IngestionPipeline(
            db=mock_db, config_loader=mock_config_loader, batch_size=100, logger=mock_logger,
        )
        pipeline._recon_repo = mock_recon_repo

        rec_date = datetime(2024, 1, 15, tzinfo=timezone.utc)
        result = await pipeline.process_file(
            file_path=str(excel_file),
            partner="MOMO",
            workflow_type="UPC",
            file_type=FileType.SETTLEMENT,
            reconciliation_date=rec_date,
        )

        assert result.file_record is not None
        assert result.file_record.processing_status == ProcessingStatus.FAILED

        events = mock_logger.events
        assert events[0][0] == "FILE_STARTED"
        failed_events = [e for e in events if e[0] == "FILE_FAILED"]
        assert len(failed_events) == 1
        assert "Config load failed" in failed_events[0][1]["error"]
