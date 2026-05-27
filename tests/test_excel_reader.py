"""Comprehensive test suite for ExcelStreamReader."""

import tempfile
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import Workbook

from src.models.mapping_config import MappingConfig
from src.readers.excel_reader import ExcelStreamReader


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def simple_excel_file(tmp_path: Path) -> Path:
    """Create a simple 2-column Excel file with headers and data."""
    path = tmp_path / "simple.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Header1", "Header2"])
    ws.append(["A", "B"])
    ws.append(["C", "D"])
    ws.append(["E", "F"])
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture
def multi_sheet_excel_file(tmp_path: Path) -> Path:
    """Create an Excel file with multiple sheets."""
    path = tmp_path / "multi_sheet.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["A", "B"])
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["X", "Y"])
    ws3 = wb.create_sheet("Sheet3")
    ws3.append(["P", "Q"])
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture
def file_with_empty_rows(tmp_path: Path) -> Path:
    """Create an Excel file with empty rows interspersed."""
    path = tmp_path / "empty_rows.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Header1", "Header2"])
    ws.append(["A", "B"])
    ws.append([None, None])  # empty row
    ws.append(["C", "D"])
    ws.append(["", ""])  # empty row (empty strings)
    ws.append(["E", "F"])
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture
def file_with_summary_rows(tmp_path: Path) -> Path:
    """Create an Excel file with summary/footer rows at bottom."""
    path = tmp_path / "summary_rows.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Header1", "Header2"])
    ws.append(["A", 100])
    ws.append(["B", 200])
    ws.append(["Total", 300])
    ws.append(["Grand Total", 300])
    ws.append(["合计", 500])
    ws.append(["C", 400])
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture
def file_with_header_row(tmp_path: Path) -> Path:
    """Create an Excel file with header on row 1, data starts row 2."""
    path = tmp_path / "header_row.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Header1", "Header2"])
    ws.append(["A", "B"])
    ws.append(["C", "D"])
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture
def large_excel_file(tmp_path: Path) -> Path:
    """Create an Excel file with 100+ rows."""
    path = tmp_path / "large.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Value"])
    for i in range(150):
        ws.append([f"row_{i}", i])
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture
def file_with_partial_rows(tmp_path: Path) -> Path:
    """Create an Excel file with partially filled rows."""
    path = tmp_path / "partial_rows.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Header1", "Header2"])
    ws.append(["A", None])  # partially filled
    ws.append([None, "B"])  # partially filled
    ws.append(["C", "D"])
    wb.save(str(path))
    wb.close()
    return path


# ---------------------------------------------------------------------------
# TestExcelStreamReaderInit
# ---------------------------------------------------------------------------

class TestExcelStreamReaderInit:
    """Tests for ExcelStreamReader initialization and validation."""

    def test_file_not_found_raises(self, tmp_path: Path) -> None:
        """FileNotFoundError raised for non-existent file."""
        missing = tmp_path / "nonexistent.xlsx"
        with pytest.raises(FileNotFoundError, match="File not found"):
            ExcelStreamReader(missing)

    def test_non_xlsx_extension_raises(self, tmp_path: Path) -> None:
        """ValueError raised for non-.xlsx/.xlsm extension."""
        csv_file = tmp_path / "data.csv"
        csv_file.write_text("a,b\n1,2")
        with pytest.raises(ValueError, match="Unsupported file extension"):
            ExcelStreamReader(csv_file)

    def test_xlsm_extension_accepted(self, tmp_path: Path) -> None:
        """.xlsm extension is accepted."""
        xlsm_file = tmp_path / "data.xlsm"
        wb = Workbook()
        wb.active.append(["A"])
        wb.save(str(xlsm_file))
        wb.close()
        reader = ExcelStreamReader(xlsm_file)
        assert reader._file_path == xlsm_file

    def test_path_object_accepted(self, simple_excel_file: Path) -> None:
        """Path object accepted as file_path."""
        reader = ExcelStreamReader(simple_excel_file)
        assert reader._file_path == simple_excel_file

    def test_string_path_accepted(self, simple_excel_file: Path) -> None:
        """String path accepted as file_path."""
        reader = ExcelStreamReader(str(simple_excel_file))
        assert reader._file_path == simple_excel_file


# ---------------------------------------------------------------------------
# TestSheetSelection
# ---------------------------------------------------------------------------

class TestSheetSelection:
    """Tests for sheet selection by name, index, and default."""

    def test_get_sheet_names(self, multi_sheet_excel_file: Path) -> None:
        """get_sheet_names returns correct list."""
        with ExcelStreamReader(multi_sheet_excel_file) as reader:
            names = reader.get_sheet_names()
            assert names == ["Sheet1", "Sheet2", "Sheet3"]

    def test_sheet_by_name(self, multi_sheet_excel_file: Path) -> None:
        """Sheet selection by name works."""
        with ExcelStreamReader(
            multi_sheet_excel_file, sheet_name="Sheet2"
        ) as reader:
            rows = list(reader.iter_rows())
            assert len(rows) == 1
            assert rows[0] == ("X", "Y")

    def test_sheet_by_name_missing_raises(self, simple_excel_file: Path) -> None:
        """KeyError raised for missing sheet name."""
        with pytest.raises(KeyError, match="Sheet 'NonExistent' not found"):
            with ExcelStreamReader(
                simple_excel_file, sheet_name="NonExistent"
            ):
                pass

    def test_sheet_by_index(self, multi_sheet_excel_file: Path) -> None:
        """Sheet selection by 0-based index works."""
        with ExcelStreamReader(
            multi_sheet_excel_file, sheet_index=1
        ) as reader:
            rows = list(reader.iter_rows())
            assert len(rows) == 1
            assert rows[0] == ("X", "Y")

    def test_sheet_by_index_out_of_range_raises(
        self, multi_sheet_excel_file: Path
    ) -> None:
        """IndexError raised for out-of-range sheet index."""
        with pytest.raises(IndexError, match="out of range"):
            with ExcelStreamReader(
                multi_sheet_excel_file, sheet_index=10
            ):
                pass

    def test_default_sheet_is_active(self, multi_sheet_excel_file: Path) -> None:
        """Default sheet is the active (first) sheet."""
        with ExcelStreamReader(multi_sheet_excel_file) as reader:
            rows = list(reader.iter_rows())
            assert len(rows) == 1
            assert rows[0] == ("A", "B")


# ---------------------------------------------------------------------------
# TestRowIteration
# ---------------------------------------------------------------------------

class TestRowIteration:
    """Tests for basic row iteration behavior."""

    def test_basic_iteration(self, simple_excel_file: Path) -> None:
        """Basic iteration yields all data rows."""
        with ExcelStreamReader(simple_excel_file, skip_empty_rows=False, skip_patterns=[]) as reader:
            rows = list(reader.iter_rows())
            assert len(rows) == 4  # header + 3 data rows
            assert rows[0] == ("Header1", "Header2")
            assert rows[1] == ("A", "B")

    def test_start_row_offset(self, file_with_header_row: Path) -> None:
        """start_row=2 skips first row."""
        with ExcelStreamReader(
            file_with_header_row, start_row=2, skip_empty_rows=False, skip_patterns=[]
        ) as reader:
            rows = list(reader.iter_rows())
            assert len(rows) == 2
            assert rows[0] == ("A", "B")
            assert rows[1] == ("C", "D")

    def test_row_tuple_values(self, simple_excel_file: Path) -> None:
        """Row tuples contain correct cell values."""
        with ExcelStreamReader(simple_excel_file, skip_empty_rows=False, skip_patterns=[]) as reader:
            rows = list(reader.iter_rows())
            assert isinstance(rows[0], tuple)
            assert rows[0] == ("Header1", "Header2")

    def test_large_file_streaming(self, large_excel_file: Path) -> None:
        """Large file streaming works (100+ rows)."""
        with ExcelStreamReader(large_excel_file, skip_empty_rows=False, skip_patterns=[]) as reader:
            rows = list(reader.iter_rows())
            assert len(rows) == 151  # header + 150 data rows

    def test_iter_rows_outside_context_raises(
        self, simple_excel_file: Path
    ) -> None:
        """iter_rows outside context manager raises RuntimeError."""
        reader = ExcelStreamReader(simple_excel_file)
        with pytest.raises(RuntimeError, match="context manager"):
            list(reader.iter_rows())


# ---------------------------------------------------------------------------
# TestEmptyRowFiltering
# ---------------------------------------------------------------------------

class TestEmptyRowFiltering:
    """Tests for empty row filtering behavior."""

    def test_empty_rows_skipped_by_default(
        self, file_with_empty_rows: Path
    ) -> None:
        """Empty rows (all None) are skipped when skip_empty_rows=True (default)."""
        with ExcelStreamReader(file_with_empty_rows) as reader:
            rows = list(reader.iter_rows())
            # header + 3 data rows (2 empty rows skipped)
            assert len(rows) == 4

    def test_empty_rows_kept_when_disabled(
        self, file_with_empty_rows: Path
    ) -> None:
        """Empty rows are kept when skip_empty_rows=False."""
        with ExcelStreamReader(
            file_with_empty_rows, skip_empty_rows=False, skip_patterns=[]
        ) as reader:
            rows = list(reader.iter_rows())
            # header + 3 data + 2 empty = 6
            assert len(rows) == 6

    def test_partially_filled_rows_not_skipped(
        self, file_with_partial_rows: Path
    ) -> None:
        """Partially filled rows are NOT skipped."""
        with ExcelStreamReader(file_with_partial_rows) as reader:
            rows = list(reader.iter_rows())
            # header + 3 data rows (none are fully empty)
            assert len(rows) == 4
            assert ("A", None) in rows
            assert (None, "B") in rows

    def test_is_empty_row_all_none(self) -> None:
        """_is_empty_row returns True for all-None row."""
        assert ExcelStreamReader._is_empty_row((None, None, None)) is True

    def test_is_empty_row_all_empty_string(self) -> None:
        """_is_empty_row returns True for all-empty-string row."""
        assert ExcelStreamReader._is_empty_row(("", "", "")) is True

    def test_is_empty_row_mixed_none_and_empty(self) -> None:
        """_is_empty_row returns True for mixed None and empty string."""
        assert ExcelStreamReader._is_empty_row((None, "", None)) is True

    def test_is_empty_row_with_data(self) -> None:
        """_is_empty_row returns False when any cell has data."""
        assert ExcelStreamReader._is_empty_row(("A", None, "")) is False

    def test_is_empty_row_with_numeric_data(self) -> None:
        """_is_empty_row returns False when cell has numeric data."""
        assert ExcelStreamReader._is_empty_row((0, None, None)) is False


# ---------------------------------------------------------------------------
# TestSummaryRowFiltering
# ---------------------------------------------------------------------------

class TestSummaryRowFiltering:
    """Tests for summary/footer row pattern filtering."""

    def test_skip_patterns_matches_case_insensitive(
        self, file_with_summary_rows: Path
    ) -> None:
        """skip_patterns matches case-insensitively."""
        with ExcelStreamReader(
            file_with_summary_rows,
            skip_empty_rows=False,
            skip_patterns=["total"],
        ) as reader:
            rows = list(reader.iter_rows())
            # "Total" and "Grand Total" should be skipped
            row_values = [r[0] for r in rows]
            assert "Total" not in row_values
            assert "Grand Total" not in row_values

    def test_skip_patterns_chinese_characters(
        self, file_with_summary_rows: Path
    ) -> None:
        """skip_patterns matches Chinese characters (合计, 总计)."""
        with ExcelStreamReader(
            file_with_summary_rows,
            skip_empty_rows=False,
            skip_patterns=["合计"],
        ) as reader:
            rows = list(reader.iter_rows())
            row_values = [r[0] for r in rows]
            assert "合计" not in row_values

    def test_default_skip_patterns(self) -> None:
        """Default skip patterns include expected keywords."""
        defaults = ExcelStreamReader.DEFAULT_SKIP_PATTERNS
        assert "total" in defaults
        assert "grand total" in defaults
        assert "summary" in defaults
        assert "footer" in defaults
        assert "合计" in defaults
        assert "总计" in defaults
        assert "小计" in defaults

    def test_skip_patterns_substring_match(self, tmp_path: Path) -> None:
        """skip_patterns matches as substring, not exact match."""
        path = tmp_path / "substring.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Description", "Amount"])
        ws.append(["Subtotal items", 100])
        ws.append(["Normal row", 200])
        wb.save(str(path))
        wb.close()
        with ExcelStreamReader(
            path, skip_empty_rows=False, skip_patterns=["total"]
        ) as reader:
            rows = list(reader.iter_rows())
            row_values = [r[0] for r in rows]
            assert "Subtotal items" not in row_values
            assert "Normal row" in row_values

    def test_empty_skip_patterns_list(self, file_with_summary_rows: Path) -> None:
        """Empty skip_patterns list means no pattern filtering."""
        with ExcelStreamReader(
            file_with_summary_rows,
            skip_empty_rows=False,
            skip_patterns=[],
        ) as reader:
            rows = list(reader.iter_rows())
            # All rows kept (no empty rows in this file)
            assert len(rows) == 7


# ---------------------------------------------------------------------------
# TestMappingConfigIntegration
# ---------------------------------------------------------------------------

class TestMappingConfigIntegration:
    """Tests for from_mapping_config factory method."""

    def test_from_mapping_config_uses_sheet_name(
        self, multi_sheet_excel_file: Path
    ) -> None:
        """from_mapping_config uses config.sheet_name."""
        config = MappingConfig(
            partner="test",
            workflowType="recon",
            fileType="SETTLEMENT",
            sheetName="Sheet2",
            startRow=1,
            fieldMappings=[],
        )
        reader = ExcelStreamReader.from_mapping_config(
            multi_sheet_excel_file, config
        )
        assert reader._sheet_name == "Sheet2"
        assert reader._start_row == 1

    def test_from_mapping_config_uses_start_row(
        self, file_with_header_row: Path
    ) -> None:
        """from_mapping_config uses config.start_row."""
        config = MappingConfig(
            partner="test",
            workflowType="recon",
            fileType="SETTLEMENT",
            sheetName="Sheet",
            startRow=2,
            fieldMappings=[],
        )
        reader = ExcelStreamReader.from_mapping_config(
            file_with_header_row, config
        )
        assert reader._start_row == 2

    def test_from_mapping_config_enables_empty_row_skip(
        self, simple_excel_file: Path
    ) -> None:
        """from_mapping_config enables skip_empty_rows by default."""
        config = MappingConfig(
            partner="test",
            workflowType="recon",
            fileType="SETTLEMENT",
            sheetName="Data",
            startRow=1,
            fieldMappings=[],
        )
        reader = ExcelStreamReader.from_mapping_config(
            simple_excel_file, config
        )
        assert reader._skip_empty_rows is True

    def test_from_mapping_config_returns_correct_type(
        self, simple_excel_file: Path
    ) -> None:
        """from_mapping_config returns ExcelStreamReader instance."""
        config = MappingConfig(
            partner="test",
            workflowType="recon",
            fileType="SETTLEMENT",
            sheetName="Data",
            startRow=1,
            fieldMappings=[],
        )
        reader = ExcelStreamReader.from_mapping_config(
            simple_excel_file, config
        )
        assert isinstance(reader, ExcelStreamReader)


# ---------------------------------------------------------------------------
# TestContextManager
# ---------------------------------------------------------------------------

class TestContextManager:
    """Tests for context manager behavior and cleanup."""

    def test_context_manager_opens_workbook(
        self, simple_excel_file: Path
    ) -> None:
        """Context manager opens workbook."""
        reader = ExcelStreamReader(simple_excel_file)
        assert reader._workbook is None
        with reader:
            assert reader._workbook is not None

    def test_context_manager_closes_workbook(
        self, simple_excel_file: Path
    ) -> None:
        """Context manager closes workbook on exit."""
        with ExcelStreamReader(simple_excel_file) as reader:
            assert reader._workbook is not None
        assert reader._workbook is None

    def test_context_manager_closes_on_exception(
        self, simple_excel_file: Path
    ) -> None:
        """Context manager closes workbook even on exception."""
        reader = ExcelStreamReader(simple_excel_file)
        with pytest.raises(ValueError):
            with reader:
                raise ValueError("test error")
        assert reader._workbook is None

    def test_get_sheet_names_outside_context_raises(
        self, simple_excel_file: Path
    ) -> None:
        """get_sheet_names outside context manager raises RuntimeError."""
        reader = ExcelStreamReader(simple_excel_file)
        with pytest.raises(RuntimeError, match="context manager"):
            reader.get_sheet_names()
