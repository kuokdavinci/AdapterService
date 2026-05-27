"""Shared pytest fixtures for ingestion pipeline integration tests.

Provides:
- mock_db: AsyncMock for AsyncIOMotorDatabase with collection mocks
- mock_config_loader: AsyncMock ConfigLoader returning sample MappingConfig
- sample_mapping_config: MappingConfig with field mappings for test data
- test_excel_file: Creates temporary .xlsx file with realistic partner data
- mock_reconciliation_file_repo: Tracks find_by_file_hash, create, update_one calls
- mock_data_container_repo: Tracks insert_many calls
"""

import tempfile
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Generator
from unittest.mock import AsyncMock, MagicMock

import openpyxl
import pytest

from src.config.loader import ConfigLoader
from src.core.enums import FileType, ProcessingStatus, TransactionStatus
from src.core.types import (
    CanonicalTransaction,
    FieldMapping,
    FieldMappingType,
    ProcessingStats,
    ValidationError,
)
from src.models.data_container import DataContainer, DataContainerRepository, PartnerData
from src.models.mapping_config import MappingConfig
from src.models.reconciliation_file import (
    ReconciliationFile,
    ReconciliationFileRepository,
)


@pytest.fixture
def mock_db() -> MagicMock:
    """AsyncMock for AsyncIOMotorDatabase with collection mocks."""
    db = MagicMock()
    # Make db[name] return a mock collection for any collection name
    db.__getitem__ = MagicMock(side_effect=lambda name: MagicMock())
    return db


@pytest.fixture
def mock_reconciliation_file_repo() -> MagicMock:
    """Mock ReconciliationFileRepository that tracks method calls."""
    repo = MagicMock(spec=ReconciliationFileRepository)
    repo.find_by_file_hash = AsyncMock(return_value=None)
    repo.create = AsyncMock(side_effect=lambda doc: doc)
    repo.update_processing_stats = AsyncMock(return_value=True)
    repo.update_status = AsyncMock(return_value=True)
    return repo


@pytest.fixture
def mock_data_container_repo() -> MagicMock:
    """Mock DataContainerRepository that tracks insert_many calls."""
    repo = MagicMock(spec=DataContainerRepository)
    repo.insert_many = AsyncMock(side_effect=lambda docs: len(docs))
    repo.find_by_duplicate_key = AsyncMock(return_value=None)
    return repo


@pytest.fixture
def sample_mapping_config() -> MappingConfig:
    """MappingConfig with field mappings for realistic partner test data.

    Maps columns:
    - A → id (STRING, required)
    - B → trace (STRING)
    - D → amount (DECIMAL, required)
    - Q → status (MAPPING with Vietnamese status values)
    - constant currency = "VND"
    """
    field_mappings = [
        FieldMapping(path="id", column="A", type=FieldMappingType.STRING, required=True),
        FieldMapping(path="trace", column="B", type=FieldMappingType.STRING),
        FieldMapping(path="amount", column="D", type=FieldMappingType.DECIMAL, required=True),
        FieldMapping(
            path="status",
            column="Q",
            type=FieldMappingType.MAPPING,
            mapping={
                "Thành công": "SUCCESS",
                "Thất bại": "FAILED",
                "Đang xử lý": "PENDING",
                "Đã hoàn tác": "REVERSED",
            },
        ),
        FieldMapping(path="currency", constant="VND", type=FieldMappingType.CONSTANT),
    ]

    return MappingConfig(
        partner="MOMO",
        workflow_type="UPC",
        file_type=FileType.SETTLEMENT,
        sheet_name="Sheet1",
        start_row=2,
        field_mappings=field_mappings,
    )


@pytest.fixture
def mock_config_loader(sample_mapping_config: MappingConfig) -> MagicMock:
    """AsyncMock ConfigLoader returning sample MappingConfig."""
    loader = MagicMock(spec=ConfigLoader)
    loader.load_by_partner_type = AsyncMock(return_value=sample_mapping_config)
    loader.load_by_version = AsyncMock(return_value=sample_mapping_config)
    return loader


def _write_row(ws, row_num: int, values: dict[int, any]) -> None:
    """Write values to specific columns in a worksheet row.

    Args:
        ws: openpyxl worksheet
        row_num: 1-based row number
        values: dict mapping column index (1-based) to value
    """
    for col_idx, value in values.items():
        ws.cell(row=row_num, column=col_idx, value=value)


# Column indices (1-based): A=1, B=2, D=4, Q=17
_COL_ID = 1
_COL_TRACE = 2
_COL_AMOUNT = 4
_COL_STATUS = 17


@pytest.fixture
def test_excel_file() -> Generator[str, None, None]:
    """Creates temporary .xlsx file with 10 realistic partner data rows.

    Columns: A (transaction ID), B (trace), D (amount), Q (status)
    Mix of valid and invalid rows:
    - Rows 1-4, 6-7, 9-10: valid rows with Vietnamese status values
    - Row 5: negative amount (invalid)
    - Row 8: empty ID (invalid)

    Yields the temp file path, cleans up after test.
    """
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Row 1: header (skipped by start_row=2)
        _write_row(ws, 1, {_COL_ID: "TransactionID", _COL_TRACE: "Trace", _COL_AMOUNT: "Amount", _COL_STATUS: "Status"})

        # Rows 2-11: data rows (10 rows total)
        # Row 2: valid — Thành công
        _write_row(ws, 2, {_COL_ID: "TXN20240115001", _COL_TRACE: "TRACE001", _COL_AMOUNT: 150000, _COL_STATUS: "Thành công"})
        # Row 3: valid — Thành công
        _write_row(ws, 3, {_COL_ID: "TXN20240115002", _COL_TRACE: "TRACE002", _COL_AMOUNT: 250000, _COL_STATUS: "Thành công"})
        # Row 4: valid — Thất bại
        _write_row(ws, 4, {_COL_ID: "TXN20240115003", _COL_TRACE: "TRACE003", _COL_AMOUNT: 75000, _COL_STATUS: "Thất bại"})
        # Row 5: valid — Thành công
        _write_row(ws, 5, {_COL_ID: "TXN20240115004", _COL_TRACE: "TRACE004", _COL_AMOUNT: 500000, _COL_STATUS: "Thành công"})
        # Row 6: INVALID — negative amount (will fail decimal validation)
        _write_row(ws, 6, {_COL_ID: "TXN20240115005", _COL_TRACE: "TRACE005", _COL_AMOUNT: -100000, _COL_STATUS: "Thành công"})
        # Row 7: valid — Đang xử lý
        _write_row(ws, 7, {_COL_ID: "TXN20240115006", _COL_TRACE: "TRACE006", _COL_AMOUNT: 320000, _COL_STATUS: "Đang xử lý"})
        # Row 8: valid — Thành công
        _write_row(ws, 8, {_COL_ID: "TXN20240115007", _COL_TRACE: "TRACE007", _COL_AMOUNT: 180000, _COL_STATUS: "Thành công"})
        # Row 9: INVALID — empty ID
        _write_row(ws, 9, {_COL_ID: "", _COL_TRACE: "TRACE008", _COL_AMOUNT: 90000, _COL_STATUS: "Thành công"})
        # Row 10: valid — Đã hoàn tác
        _write_row(ws, 10, {_COL_ID: "TXN20240115009", _COL_TRACE: "TRACE009", _COL_AMOUNT: 420000, _COL_STATUS: "Đã hoàn tác"})
        # Row 11: valid — Thành công
        _write_row(ws, 11, {_COL_ID: "TXN20240115010", _COL_TRACE: "TRACE010", _COL_AMOUNT: 610000, _COL_STATUS: "Thành công"})

        wb.save(f.name)
        temp_path = f.name

    yield temp_path

    # Cleanup
    Path(temp_path).unlink(missing_ok=True)


@pytest.fixture
def empty_excel_file() -> Generator[str, None, None]:
    """Creates temporary .xlsx file with header only (no data rows).

    Yields the temp file path, cleans up after test.
    """
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # Row 1: header only (skipped by start_row=2)
        _write_row(ws, 1, {_COL_ID: "TransactionID", _COL_TRACE: "Trace", _COL_AMOUNT: "Amount", _COL_STATUS: "Status"})
        wb.save(f.name)
        temp_path = f.name

    yield temp_path

    Path(temp_path).unlink(missing_ok=True)


@pytest.fixture
def all_invalid_excel_file() -> Generator[str, None, None]:
    """Creates temporary .xlsx file where ALL data rows are invalid.

    Yields the temp file path, cleans up after test.
    """
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Row 1: header
        _write_row(ws, 1, {_COL_ID: "TransactionID", _COL_TRACE: "Trace", _COL_AMOUNT: "Amount", _COL_STATUS: "Status"})

        # All rows invalid: empty IDs
        _write_row(ws, 2, {_COL_ID: "", _COL_TRACE: "TRACE001", _COL_AMOUNT: 100000, _COL_STATUS: "Thành công"})
        _write_row(ws, 3, {_COL_ID: "", _COL_TRACE: "TRACE002", _COL_AMOUNT: 200000, _COL_STATUS: "Thành công"})
        _write_row(ws, 4, {_COL_ID: "", _COL_TRACE: "TRACE003", _COL_AMOUNT: 300000, _COL_STATUS: "Thành công"})

        wb.save(f.name)
        temp_path = f.name

    yield temp_path

    Path(temp_path).unlink(missing_ok=True)


@pytest.fixture
def large_excel_file() -> Generator[str, None, None]:
    """Creates temporary .xlsx file with 250 valid rows for batch testing.

    Yields the temp file path, cleans up after test.
    """
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Row 1: header
        _write_row(ws, 1, {_COL_ID: "TransactionID", _COL_TRACE: "Trace", _COL_AMOUNT: "Amount", _COL_STATUS: "Status"})

        # 250 data rows
        for i in range(1, 251):
            _write_row(ws, 1 + i, {
                _COL_ID: f"TXN20240115{i:04d}",
                _COL_TRACE: f"TRACE{i:04d}",
                _COL_AMOUNT: 100000 + i * 100,
                _COL_STATUS: "Thành công",
            })

        wb.save(f.name)
        temp_path = f.name

    yield temp_path

    Path(temp_path).unlink(missing_ok=True)
